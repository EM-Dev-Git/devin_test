"""
コーディング規約のExcelドキュメントを更新するスクリプト

このスクリプトは既存のコーディング規約Excelファイルを読み込み、
プロジェクト構成のディレクトリ構造を複数のセルに分割して記載します。
"""
import os
import sys
import openpyxl
from openpyxl.styles import Font, Alignment

def update_project_structure_sheet(excel_file):
    """
    プロジェクト構成シートのディレクトリ構造を複数セルに分割する
    
    Args:
        excel_file: 更新するExcelファイルのパス
        
    Returns:
        bool: 更新が成功した場合はTrue、それ以外はFalse
    """
    try:
        wb = openpyxl.load_workbook(excel_file)
        
        ws = wb["1_プロジェクト構成"]
        
        current_structure = ws['C5'].value
        
        ws['C5'].value = None
        
        structure_lines = current_structure.strip().split('\n')
        
        for i, line in enumerate(structure_lines):
            row = 5 + i
            ws[f'C{row}'] = line
            ws[f'C{row}'].font = Font(name='Consolas', size=10)
            ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        explanation_start_row = 5 + len(structure_lines) + 2
        
        ws[f'B{explanation_start_row}'] = "ディレクトリの役割"
        ws[f'B{explanation_start_row}'].font = Font(name='Yu Gothic', size=12, bold=True)
        
        for row in range(7, 14):
            ws[f'B{row}'].value = None
            ws[f'C{row}'].value = None
        
        directories = [
            ("api/", "APIエンドポイントとルーティングを定義します。バージョン管理のためにv1などのサブディレクトリを使用します。"),
            ("core/", "アプリケーションの中核となる設定や初期化処理を配置します。"),
            ("models/", "データベースモデル（SQLAlchemy）とスキーマ（Pydantic）を定義します。"),
            ("services/", "ビジネスロジックを実装します。APIとデータベース層の間の処理を担当します。"),
            ("db/", "データベース接続やセッション管理に関するコードを配置します。"),
            ("dependencies/", "FastAPIの依存関係（Depends）を定義します。認証やデータベースセッションなど。")
        ]
        
        for i, (dir_name, description) in enumerate(directories):
            row = explanation_start_row + i + 1
            ws[f'B{row}'] = dir_name
            ws[f'B{row}'].font = Font(name='Consolas', size=11, bold=True)
            ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
            
            ws[f'C{row}'] = description
            ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
            ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        wb.save(excel_file)
        print(f"Successfully updated project structure in {excel_file}")
        return True
        
    except Exception as e:
        print(f"Error updating project structure: {str(e)}")
        return False

def main():
    """
    メイン関数
    """
    coding_standards_file = "DOC/コーディング規約/コーディング規約.xlsx"
    
    if not os.path.exists(coding_standards_file):
        print(f"Error: Coding standards file not found at {coding_standards_file}")
        return 1
        
    success = update_project_structure_sheet(coding_standards_file)
    
    if success:
        print("Excel update successful.")
        return 0
    else:
        print("Excel update failed.")
        return 1

if __name__ == "__main__":
    sys.exit(main())
