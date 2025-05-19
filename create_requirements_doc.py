import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

wb = openpyxl.Workbook()

sheets = [
    '表紙', '改訂履歴', '機能要件', '非機能要件', 
    'API要件', 'データ要件', '統合要件'
]

wb.remove(wb.active)
for sheet_name in sheets:
    wb.create_sheet(sheet_name)

title_font = Font(name='Yu Gothic', size=16, bold=True)
header_font = Font(name='Yu Gothic', size=12, bold=True)
normal_font = Font(name='Yu Gothic', size=11)
header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

cover = wb['表紙']
cover['B2'] = 'OpenAI連携機能 要件定義書'
cover['B2'].font = Font(name='Yu Gothic', size=20, bold=True)
cover['B4'] = 'プロジェクト名：'
cover['C4'] = 'EM_test_project'
cover['B5'] = '作成日：'
cover['C5'] = datetime.now().strftime('%Y年%m月%d日')
cover['B6'] = '作成者：'
cover['C6'] = 'Devin AI'
cover['B8'] = '概要：'
cover['C8'] = 'このドキュメントはOpenAI APIを使用したAI問い合わせ機能の要件を定義します。'
cover['C9'] = 'ユーザーからの入力に基づいてAIレスポンスを生成し、結果を返す機能を実装します。'

for col in range(1, 10):
    cover.column_dimensions[get_column_letter(col)].width = 15

history = wb['改訂履歴']
history['A1'] = '改訂履歴'
history['A1'].font = title_font
history['A3'] = 'バージョン'
history['B3'] = '日付'
history['C3'] = '作成者'
history['D3'] = '変更内容'
history['E3'] = '承認者'

for cell in history[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

history['A4'] = '1.0'
history['B4'] = datetime.now().strftime('%Y/%m/%d')
history['C4'] = 'Devin AI'
history['D4'] = '初版作成'
history['E4'] = ''

for cell in history[4]:
    cell.font = normal_font
    cell.border = border

history.column_dimensions['A'].width = 12
history.column_dimensions['B'].width = 12
history.column_dimensions['C'].width = 15
history.column_dimensions['D'].width = 40
history.column_dimensions['E'].width = 15

func = wb['機能要件']
func['A1'] = 'OpenAI連携機能 機能要件'
func['A1'].font = title_font

func['A3'] = 'ID'
func['B3'] = '要件名'
func['C3'] = '説明'
func['D3'] = '優先度'
func['E3'] = '状態'

for cell in func[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

requirements = [
    ('FR-001', 'AIチャット機能', 'ユーザーからの入力テキストに基づいてAIレスポンスを生成する', '高', '計画済'),
    ('FR-002', 'プロンプトテンプレート', 'AIへの入力に使用するプロンプトテンプレートを定義する', '高', '計画済'),
    ('FR-003', 'レスポンスパラメータ制御', 'トークン数や温度などのAIレスポンス生成パラメータを制御する', '中', '計画済'),
    ('FR-004', 'ユーザー認証', 'AIチャット機能へのアクセスにはユーザー認証が必要', '高', '計画済'),
    ('FR-005', 'エラーハンドリング', 'API接続エラーや無効な入力に対する適切なエラーハンドリング', '中', '計画済'),
    ('FR-006', 'ログ記録', 'AIリクエストとレスポンスの詳細をログに記録する', '中', '計画済'),
]

for i, req in enumerate(requirements, 4):
    func[f'A{i}'] = req[0]
    func[f'B{i}'] = req[1]
    func[f'C{i}'] = req[2]
    func[f'D{i}'] = req[3]
    func[f'E{i}'] = req[4]
    
    for col in range(5):
        cell = func[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col == 2:  # Description column
            cell.alignment = Alignment(wrap_text=True)

func.column_dimensions['A'].width = 10
func.column_dimensions['B'].width = 20
func.column_dimensions['C'].width = 50
func.column_dimensions['D'].width = 10
func.column_dimensions['E'].width = 10

non_func = wb['非機能要件']
non_func['A1'] = 'OpenAI連携機能 非機能要件'
non_func['A1'].font = title_font

non_func['A3'] = 'ID'
non_func['B3'] = '要件名'
non_func['C3'] = '説明'
non_func['D3'] = '優先度'
non_func['E3'] = '状態'

for cell in non_func[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

requirements = [
    ('NF-001', 'パフォーマンス', 'AIレスポンス生成は5秒以内に完了すること', '高', '計画済'),
    ('NF-002', 'セキュリティ', 'OpenAI APIキーは環境変数で安全に管理すること', '高', '計画済'),
    ('NF-003', 'スケーラビリティ', '同時に複数のリクエストを処理できること', '中', '計画済'),
    ('NF-004', '可用性', 'APIエラー発生時も適切なエラーメッセージを返し、アプリケーションがクラッシュしないこと', '高', '計画済'),
    ('NF-005', '保守性', 'プロンプトテンプレートは別ファイルで管理し、容易に変更できること', '中', '計画済'),
]

for i, req in enumerate(requirements, 4):
    non_func[f'A{i}'] = req[0]
    non_func[f'B{i}'] = req[1]
    non_func[f'C{i}'] = req[2]
    non_func[f'D{i}'] = req[3]
    non_func[f'E{i}'] = req[4]
    
    for col in range(5):
        cell = non_func[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col == 2:  # Description column
            cell.alignment = Alignment(wrap_text=True)

non_func.column_dimensions['A'].width = 10
non_func.column_dimensions['B'].width = 20
non_func.column_dimensions['C'].width = 50
non_func.column_dimensions['D'].width = 10
non_func.column_dimensions['E'].width = 10

api = wb['API要件']
api['A1'] = 'OpenAI連携機能 API要件'
api['A1'].font = title_font

api['A3'] = 'ID'
api['B3'] = 'エンドポイント'
api['C3'] = '説明'
api['D3'] = 'リクエスト'
api['E3'] = 'レスポンス'
api['F3'] = '認証'

for cell in api[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

requirements = [
    ('API-001', '/api/v1/llm/chat', 'AIとチャットするためのエンドポイント', 'prompt, max_tokens, temperature', 'response, model, usage', '必須'),
]

for i, req in enumerate(requirements, 4):
    api[f'A{i}'] = req[0]
    api[f'B{i}'] = req[1]
    api[f'C{i}'] = req[2]
    api[f'D{i}'] = req[3]
    api[f'E{i}'] = req[4]
    api[f'F{i}'] = req[5]
    
    for col in range(6):
        cell = api[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 3, 4]:  # Description, request, response columns
            cell.alignment = Alignment(wrap_text=True)

api.column_dimensions['A'].width = 10
api.column_dimensions['B'].width = 15
api.column_dimensions['C'].width = 30
api.column_dimensions['D'].width = 25
api.column_dimensions['E'].width = 25
api.column_dimensions['F'].width = 10

data = wb['データ要件']
data['A1'] = 'OpenAI連携機能 データ要件'
data['A1'].font = title_font

data['A3'] = 'ID'
data['B3'] = 'データモデル'
data['C3'] = '説明'
data['D3'] = '属性'
data['E3'] = '制約'

for cell in data[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

requirements = [
    ('DT-001', 'LLMRequest', 'AIリクエストのデータモデル', 'prompt, max_tokens, temperature', 'promptは必須'),
    ('DT-002', 'LLMResponse', 'AIレスポンスのデータモデル', 'response, model, usage', 'responseは必須'),
]

for i, req in enumerate(requirements, 4):
    data[f'A{i}'] = req[0]
    data[f'B{i}'] = req[1]
    data[f'C{i}'] = req[2]
    data[f'D{i}'] = req[3]
    data[f'E{i}'] = req[4]
    
    for col in range(5):
        cell = data[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 3]:  # Description, attributes columns
            cell.alignment = Alignment(wrap_text=True)

data.column_dimensions['A'].width = 10
data.column_dimensions['B'].width = 15
data.column_dimensions['C'].width = 30
data.column_dimensions['D'].width = 25
data.column_dimensions['E'].width = 15

integration = wb['統合要件']
integration['A1'] = 'OpenAI連携機能 統合要件'
integration['A1'].font = title_font

integration['A3'] = 'ID'
integration['B3'] = '要件名'
integration['C3'] = '説明'
integration['D3'] = '依存関係'
integration['E3'] = '状態'

for cell in integration[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

requirements = [
    ('INT-001', 'OpenAI API連携', 'OpenAI APIと連携してAIレスポンスを生成する', 'OpenAI APIキー', '計画済'),
    ('INT-002', '認証システム連携', 'ユーザー認証システムと連携してAPIアクセスを制御する', 'JWT認証', '計画済'),
    ('INT-003', 'ロギングシステム連携', 'アプリケーションのロギングシステムと連携してリクエスト/レスポンスをログに記録する', 'ロギングシステム', '計画済'),
]

for i, req in enumerate(requirements, 4):
    integration[f'A{i}'] = req[0]
    integration[f'B{i}'] = req[1]
    integration[f'C{i}'] = req[2]
    integration[f'D{i}'] = req[3]
    integration[f'E{i}'] = req[4]
    
    for col in range(5):
        cell = integration[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col == 2:  # Description column
            cell.alignment = Alignment(wrap_text=True)

integration.column_dimensions['A'].width = 10
integration.column_dimensions['B'].width = 20
integration.column_dimensions['C'].width = 50
integration.column_dimensions['D'].width = 15
integration.column_dimensions['E'].width = 10

os.makedirs('DOC/要件定義', exist_ok=True)

wb.save('DOC/要件定義/OpenAI連携機能_要件定義書.xlsx')
print('Excel file created successfully: DOC/要件定義/OpenAI連携機能_要件定義書.xlsx')
