"""
コーディング規約のExcelドキュメントを生成するスクリプト（マルチシート版）

このスクリプトはopenpyxlライブラリを使用して、
FastAPIプロジェクトのコーディング規約をExcel形式で生成します。
表紙、改訂履歴、目次、各セクションを別々のシートに分けて作成します。
"""
import os
import sys
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

def create_cover_sheet(wb):
    """表紙シートを作成する"""
    ws = wb.create_sheet("表紙", 0)
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    ws.merge_cells('B2:F2')
    ws['B2'] = "FastAPI プロジェクト"
    ws['B2'].font = Font(name='Yu Gothic', size=16, bold=True)
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('B3:F3')
    ws['B3'] = "コーディング規約"
    ws['B3'].font = Font(name='Yu Gothic', size=24, bold=True)
    ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('B5:C5')
    ws['B5'] = "作成日:"
    ws['B5'].font = Font(name='Yu Gothic', size=12, bold=True)
    ws['B5'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws.merge_cells('D5:F5')
    ws['D5'] = datetime.datetime.now().strftime("%Y年%m月%d日")
    ws['D5'].font = Font(name='Yu Gothic', size=12)
    ws['D5'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws.merge_cells('B6:C6')
    ws['B6'] = "バージョン:"
    ws['B6'].font = Font(name='Yu Gothic', size=12, bold=True)
    ws['B6'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws.merge_cells('D6:F6')
    ws['D6'] = "1.0"
    ws['D6'].font = Font(name='Yu Gothic', size=12)
    ws['D6'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws.merge_cells('B7:C7')
    ws['B7'] = "作成者:"
    ws['B7'].font = Font(name='Yu Gothic', size=12, bold=True)
    ws['B7'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws.merge_cells('D7:F7')
    ws['D7'] = "EM開発チーム"
    ws['D7'].font = Font(name='Yu Gothic', size=12)
    ws['D7'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws.merge_cells('B9:F9')
    ws['B9'] = "目的"
    ws['B9'].font = Font(name='Yu Gothic', size=14, bold=True)
    ws['B9'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws.merge_cells('B10:F12')
    ws['B10'] = "このドキュメントはFastAPIプロジェクトのコーディング規約を定義し、チーム開発やコードの一貫性を保つことを目的としています。すべての開発者はこの規約に従ってコードを作成することが求められます。"
    ws['B10'].font = Font(name='Yu Gothic', size=11)
    ws['B10'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 40
    ws.row_dimensions[10].height = 60
    
    return ws

def create_revision_history_sheet(wb):
    """改訂履歴シートを作成する"""
    ws = wb.create_sheet("改訂履歴", 1)
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    
    ws.merge_cells('B2:E2')
    ws['B2'] = "改訂履歴"
    ws['B2'].font = Font(name='Yu Gothic', size=16, bold=True)
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    
    headers = ["バージョン", "改訂日", "改訂内容", "改訂者"]
    cols = ['B', 'C', 'D', 'E']
    
    for i, header in enumerate(headers):
        ws[f'{cols[i]}4'] = header
        ws[f'{cols[i]}4'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'{cols[i]}4'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{cols[i]}4'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    ws['B5'] = "1.0"
    ws['C5'] = datetime.datetime.now().strftime("%Y/%m/%d")
    ws['D5'] = "初版作成"
    ws['E5'] = "EM開発チーム"
    
    for col in cols:
        ws[f'{col}5'].font = Font(name='Yu Gothic', size=11)
        ws[f'{col}5'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[4].height = 20
    
    return ws

def create_toc_sheet(wb):
    """目次シートを作成する"""
    ws = wb.create_sheet("目次", 2)
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    
    ws.merge_cells('B2:D2')
    ws['B2'] = "目次"
    ws['B2'].font = Font(name='Yu Gothic', size=16, bold=True)
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['B4'] = "項番"
    ws['C4'] = "項目名"
    ws['D4'] = "シート名"
    
    for col in ['B', 'C', 'D']:
        ws[f'{col}4'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'{col}4'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}4'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    toc_items = [
        ("1", "プロジェクト構成", "1_プロジェクト構成"),
        ("2", "命名規則", "2_命名規則"),
        ("3", "型ヒントとドキュメンテーション", "3_型ヒント"),
        ("4", "Pydanticモデルの使い方", "4_Pydanticモデル"),
        ("5", "依存関係の注入", "5_依存関係"),
        ("6", "設定ファイル", "6_設定ファイル"),
        ("7", "ログ出力の統一", "7_ログ出力"),
        ("8", "テストの命名と構成", "8_テスト"),
        ("9", "推奨ツール", "9_推奨ツール")
    ]
    
    for i, (num, name, sheet) in enumerate(toc_items):
        row = i + 5
        ws[f'B{row}'] = num
        ws[f'C{row}'] = name
        ws[f'D{row}'] = sheet
        
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=11)
        ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
        ws[f'D{row}'].font = Font(name='Yu Gothic', size=11)
        
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[4].height = 20
    
    return ws

def create_project_structure_sheet(wb):
    """プロジェクト構成シートを作成する"""
    ws = wb.create_sheet("1_プロジェクト構成", 3)
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    
    ws.merge_cells('B2:C2')
    ws['B2'] = "1. プロジェクト構成"
    ws['B2'].font = Font(name='Yu Gothic', size=16, bold=True)
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws.merge_cells('B3:C3')
    ws['B3'] = "FastAPIプロジェクトの標準的なディレクトリ構成を以下に示します。この構成に従うことで、コードの可読性と保守性が向上します。"
    ws['B3'].font = Font(name='Yu Gothic', size=11)
    ws['B3'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws['B5'] = "ディレクトリ構成"
    ws['B5'].font = Font(name='Yu Gothic', size=12, bold=True)
    
    structure = """
app/
├── api/              # ルーティング関連
│   ├── v1/
│   │   ├── endpoints/
│   │   │   └── user.py
│   │   └── router.py
├── core/             # 設定や初期化処理
│   ├── config.py
│   └── security.py
├── models/           # PydanticモデルやORMモデル
│   └── user.py
├── services/         # ビジネスロジック
│   └── user_service.py
├── db/               # DB接続やセッション管理
│   ├── base.py
│   └── session.py
├── main.py           # アプリケーションのエントリーポイント
└── dependencies/     # 共通依存関係
    """
    
    ws['C5'] = structure
    ws['C5'].font = Font(name='Consolas', size=10)
    ws['C5'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws['B7'] = "ディレクトリの役割"
    ws['B7'].font = Font(name='Yu Gothic', size=12, bold=True)
    
    directories = [
        ("api/", "APIエンドポイントとルーティングを定義します。バージョン管理のためにv1などのサブディレクトリを使用します。"),
        ("core/", "アプリケーションの中核となる設定や初期化処理を配置します。"),
        ("models/", "データベースモデル（SQLAlchemy）とスキーマ（Pydantic）を定義します。"),
        ("services/", "ビジネスロジックを実装します。APIとデータベース層の間の処理を担当します。"),
        ("db/", "データベース接続やセッション管理に関するコードを配置します。"),
        ("dependencies/", "FastAPIの依存関係（Depends）を定義します。認証やデータベースセッションなど。")
    ]
    
    for i, (dir_name, description) in enumerate(directories):
        row = i + 8
        ws[f'B{row}'] = dir_name
        ws[f'B{row}'].font = Font(name='Consolas', size=11, bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws[f'C{row}'] = description
        ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
        ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 40
    ws.row_dimensions[5].height = 200
    
    return ws

def create_naming_conventions_sheet(wb):
    """命名規則シートを作成する"""
    ws = wb.create_sheet("2_命名規則", 4)
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    
    ws.merge_cells('B2:D2')
    ws['B2'] = "2. 命名規則"
    ws['B2'].font = Font(name='Yu Gothic', size=16, bold=True)
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws.merge_cells('B3:D3')
    ws['B3'] = "一貫性のある命名規則を使用することで、コードの可読性と保守性が向上します。以下の命名規則に従ってください。"
    ws['B3'].font = Font(name='Yu Gothic', size=11)
    ws['B3'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws['B5'] = "項目"
    ws['C5'] = "規則"
    ws['D5'] = "例"
    
    for col in ['B', 'C', 'D']:
        ws[f'{col}5'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'{col}5'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}5'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    naming_rules = [
        ("ファイル名", "snake_case", "user_service.py"),
        ("クラス名", "PascalCase", "UserService"),
        ("関数・変数名", "snake_case", "get_user_by_id"),
        ("定数", "UPPER_CASE", "MAX_CONNECTIONS"),
        ("モジュール名", "snake_case", "auth_utils"),
        ("パッケージ名", "snake_case", "api_client"),
        ("エンドポイント", "kebab-case", "/api/v1/user-profiles"),
        ("データベーステーブル", "snake_case、複数形", "user_profiles")
    ]
    
    for i, (item, rule, example) in enumerate(naming_rules):
        row = i + 6
        ws[f'B{row}'] = item
        ws[f'C{row}'] = rule
        ws[f'D{row}'] = example
        
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=11)
        ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
        ws[f'D{row}'].font = Font(name='Consolas', size=10)
        
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'D{row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws.merge_cells('B15:D15')
    ws['B15'] = "命名のベストプラクティス"
    ws['B15'].font = Font(name='Yu Gothic', size=12, bold=True)
    ws['B15'].alignment = Alignment(horizontal='left', vertical='center')
    
    best_practices = [
        "• 意味のある名前を使用する（x, y, tempなどの意味のない名前は避ける）",
        "• 略語は一般的に知られているもの以外は避ける",
        "• 関数名は動詞から始める（get_user, create_item）",
        "• ブール値を返す関数は is_, has_, can_ などから始める（is_active, has_permission）",
        "• 一貫性を保つ（同じ概念には同じ命名パターンを使用する）"
    ]
    
    for i, practice in enumerate(best_practices):
        row = i + 16
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'] = practice
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=11)
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 40
    ws.row_dimensions[15].height = 25
    
    return ws

def create_remaining_sheets(wb):
    """残りのシートを作成する"""
    ws = wb.create_sheet("3_型ヒント", 5)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    
    ws.merge_cells('B2:C2')
    ws['B2'] = "3. 型ヒントとドキュメンテーション"
    ws['B2'].font = Font(name='Yu Gothic', size=16, bold=True)
    
    ws.merge_cells('B3:C3')
    ws['B3'] = "型ヒントとドキュメンテーションを使用することで、コードの理解しやすさと保守性が向上します。"
    ws['B3'].font = Font(name='Yu Gothic', size=11)
    ws['B3'].alignment = Alignment(wrap_text=True)
    
    ws['B5'] = "型ヒントの基本"
    ws['B5'].font = Font(name='Yu Gothic', size=12, bold=True)
    
    ws['C5'] = """
def get_user_by_id(user_id: int) -> dict:
    \"\"\"ユーザーIDからユーザー情報を取得する\"\"\"
    """
    ws['C5'].font = Font(name='Consolas', size=10)
    ws['C5'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws = wb.create_sheet("4_Pydanticモデル", 6)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    
    ws.merge_cells('B2:C2')
    ws['B2'] = "4. Pydanticモデルの使い方"
    ws['B2'].font = Font(name='Yu Gothic', size=16, bold=True)
    
    ws.merge_cells('B3:C3')
    ws['B3'] = "Pydanticモデルを使用して、APIのリクエストとレスポンスのデータ構造を定義します。"
    ws['B3'].font = Font(name='Yu Gothic', size=11)
    ws['B3'].alignment = Alignment(wrap_text=True)
    
    ws['B5'] = "基本的なモデル"
    ws['B5'].font = Font(name='Yu Gothic', size=12, bold=True)
    
    ws['C5'] = """
from pydantic import BaseModel, EmailStr

class UserCreate(BaseModel):
    email: EmailStr
    password: str

    class Config:
        orm_mode = True
    """
    ws['C5'].font = Font(name='Consolas', size=10)
    ws['C5'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws = wb.create_sheet("5_依存関係", 7)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    
    ws.merge_cells('B2:C2')
    ws['B2'] = "5. 依存関係の注入"
    ws['B2'].font = Font(name='Yu Gothic', size=16, bold=True)
    
    ws.merge_cells('B3:C3')
    ws['B3'] = "FastAPIの依存関係注入システムを使用して、コードの再利用性と保守性を向上させます。"
    ws['B3'].font = Font(name='Yu Gothic', size=11)
    ws['B3'].alignment = Alignment(wrap_text=True)
    
    ws['B5'] = "基本的な依存関係"
    ws['B5'].font = Font(name='Yu Gothic', size=12, bold=True)
    
    ws['C5'] = """
from fastapi import Depends
from sqlalchemy.orm import Session
from app.db.session import get_db

@router.post("/users")
def create_user(user: UserCreate, db: Session = Depends(get_db)):
    return user_service.create_user(db, user)
    """
    ws['C5'].font = Font(name='Consolas', size=10)
    ws['C5'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws = wb.create_sheet("6_設定ファイル", 8)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    
    ws.merge_cells('B2:C2')
    ws['B2'] = "6. 設定ファイル"
    ws['B2'].font = Font(name='Yu Gothic', size=16, bold=True)
    
    ws.merge_cells('B3:C3')
    ws['B3'] = "アプリケーションの設定は環境変数から読み込み、Pydanticモデルを使用して型安全に管理します。"
    ws['B3'].font = Font(name='Yu Gothic', size=11)
    ws['B3'].alignment = Alignment(wrap_text=True)
    
    ws['B5'] = "設定クラス"
    ws['B5'].font = Font(name='Yu Gothic', size=12, bold=True)
    
    ws['C5'] = """
from pydantic import BaseSettings

class Settings(BaseSettings):
    app_name: str = "My FastAPI App"
    database_url: str

    class Config:
        env_file = ".env"

settings = Settings()
    """
    ws['C5'].font = Font(name='Consolas', size=10)
    ws['C5'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws = wb.create_sheet("7_ログ出力", 9)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    
    ws.merge_cells('B2:C2')
    ws['B2'] = "7. ログ出力の統一"
    ws['B2'].font = Font(name='Yu Gothic', size=16, bold=True)
    
    ws.merge_cells('B3:C3')
    ws['B3'] = "ログ出力を統一することで、アプリケーションの動作を追跡しやすくなります。"
    ws['B3'].font = Font(name='Yu Gothic', size=11)
    ws['B3'].alignment = Alignment(wrap_text=True)
    
    ws['B5'] = "ロガーの設定"
    ws['B5'].font = Font(name='Yu Gothic', size=12, bold=True)
    
    ws['C5'] = """
import logging

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
    """
    ws['C5'].font = Font(name='Consolas', size=10)
    ws['C5'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws = wb.create_sheet("8_テスト", 10)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    
    ws.merge_cells('B2:C2')
    ws['B2'] = "8. テストの命名と構成"
    ws['B2'].font = Font(name='Yu Gothic', size=16, bold=True)
    
    ws.merge_cells('B3:C3')
    ws['B3'] = "テストコードは一貫した命名規則と構成に従うことで、保守性と理解しやすさが向上します。"
    ws['B3'].font = Font(name='Yu Gothic', size=11)
    ws['B3'].alignment = Alignment(wrap_text=True)
    
    ws['B5'] = "テスト関数"
    ws['B5'].font = Font(name='Yu Gothic', size=12, bold=True)
    
    ws['C5'] = """
def test_create_user(client):
    response = client.post("/users", json={"email": "test@example.com", "password": "123456"})
    assert response.status_code == 200
    """
    ws['C5'].font = Font(name='Consolas', size=10)
    ws['C5'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws = wb.create_sheet("9_推奨ツール", 11)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    
    ws.merge_cells('B2:C2')
    ws['B2'] = "9. 推奨ツール"
    ws['B2'].font = Font(name='Yu Gothic', size=16, bold=True)
    
    ws.merge_cells('B3:C3')
    ws['B3'] = "以下のツールを使用することで、コードの品質と一貫性を保つことができます。"
    ws['B3'].font = Font(name='Yu Gothic', size=11)
    ws['B3'].alignment = Alignment(wrap_text=True)
    
    tools = [
        ("Black", "コードフォーマッター", "コードスタイルを自動的に統一します。"),
        ("isort", "インポート整理", "インポート文を整理します。"),
        ("mypy", "静的型チェック", "型ヒントに基づいて型エラーを検出します。"),
        ("ruff", "リンター", "コードの問題を検出します。")
    ]
    
    ws['B5'] = "ツール名"
    ws['C5'] = "説明"
    
    for col in ['B', 'C']:
        ws[f'{col}5'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'{col}5'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}5'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for i, (tool, category, description) in enumerate(tools):
        row = i + 6
        ws[f'B{row}'] = tool
        ws[f'C{row}'] = f"{category}: {description}"
        
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
        
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    return wb

def create_coding_standards_doc(output_file):
    """
    コーディング規約のExcelドキュメントを生成する
    
    Args:
        output_file: 出力Excelファイルのパス
        
    Returns:
        bool: 生成が成功した場合はTrue、それ以外はFalse
    """
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        create_cover_sheet(wb)
        create_revision_history_sheet(wb)
        create_toc_sheet(wb)
        create_project_structure_sheet(wb)
        create_naming_conventions_sheet(wb)
        create_remaining_sheets(wb)
        
        wb.save(output_file)
        print(f"Successfully created coding standards document at {output_file}")
        return True
        
    except Exception as e:
        print(f"Error creating coding standards document: {str(e)}")
        return False

def main():
    """
    メイン関数
    """
    coding_standards_dir = "DOC/コーディング規約"
    if not os.path.exists(coding_standards_dir):
        os.makedirs(coding_standards_dir)
        
    readme_path = os.path.join(coding_standards_dir, "README.md")
    with open(readme_path, "w", encoding="utf-8") as f:
        f.write("""# コーディング規約

このディレクトリにはFastAPIプロジェクトのコーディング規約に関するドキュメントが含まれています。


- `コーディング規約.xlsx`: FastAPIプロジェクトのコーディング規約を定義したExcelファイル


1. **表紙**: ドキュメントのタイトル、バージョン、作成日、作成者
2. **改訂履歴**: ドキュメントの変更履歴
3. **目次**: 各セクションへのリンク
4. **各セクション**: プロジェクト構成、命名規則、型ヒント、Pydanticモデルなど

各セクションは別々のシートに分かれており、内容は複数のセルに分散して記載されています。
""")
    
    output_file = os.path.join(coding_standards_dir, "コーディング規約.xlsx")
    success = create_coding_standards_doc(output_file)
    
    if success:
        print(f"Excel generation successful. Excel file saved to {output_file}")
        return 0
    else:
        print("Excel generation failed")
        return 1

if __name__ == "__main__":
    sys.exit(main())
