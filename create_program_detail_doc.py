import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

wb = openpyxl.Workbook()

sheets = [
    '表紙', '改訂履歴', 'モジュール構成', 'ファイル一覧', 
    '関数仕様', 'クラス実装', 'データベース', 'エラー処理', 'テスト仕様'
]

wb.remove(wb.active)
for sheet_name in sheets:
    wb.create_sheet(sheet_name)

title_font = Font(name='Yu Gothic', size=16, bold=True)
header_font = Font(name='Yu Gothic', size=12, bold=True)
normal_font = Font(name='Yu Gothic', size=11)
header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

cover = wb['表紙']
cover['B2'] = 'OpenAI連携機能 プログラム詳細設計書'
cover['B2'].font = Font(name='Yu Gothic', size=20, bold=True)
cover['B4'] = 'プロジェクト名：'
cover['C4'] = 'EM_test_project'
cover['B5'] = '作成日：'
cover['C5'] = datetime.now().strftime('%Y年%m月%d日')
cover['B6'] = '作成者：'
cover['C6'] = 'Devin AI'
cover['B8'] = '概要：'
cover['C8'] = 'このドキュメントはOpenAI APIを使用したAI問い合わせ機能のプログラム詳細設計を定義します。'
cover['C9'] = 'モジュール構成、ファイル一覧、関数仕様、クラス実装、データベース設計、エラー処理、テスト仕様を含みます。'

for col in range(1, 10):
    cover.column_dimensions[get_column_letter(col)].width = 15

history = wb['改訂履歴']
history['A1'] = '改訂履歴'
history['A1'].font = title_font
history['A3'] = 'バージョン'
history['B3'] = '日付'
history['C3'] = '作成者'
history['D3'] = '変更内容'
history['E3'] = '承認者'

for cell in history[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

history['A4'] = '1.0'
history['B4'] = datetime.now().strftime('%Y/%m/%d')
history['C4'] = 'Devin AI'
history['D4'] = '初版作成'
history['E4'] = ''

for cell in history[4]:
    cell.font = normal_font
    cell.border = border

history.column_dimensions['A'].width = 12
history.column_dimensions['B'].width = 12
history.column_dimensions['C'].width = 15
history.column_dimensions['D'].width = 40
history.column_dimensions['E'].width = 15

module = wb['モジュール構成']
module['A1'] = 'OpenAI連携機能 モジュール構成'
module['A1'].font = title_font

module['A3'] = 'ID'
module['B3'] = 'モジュール名'
module['C3'] = '説明'
module['D3'] = '依存モジュール'
module['E3'] = '責任者'

for cell in module[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

modules = [
    ('MOD-001', 'app.routes.llm', 'OpenAI APIエンドポイントを提供するルーターモジュール', 'app.utils.openai_client, app.prompts.chat_prompt', 'AI開発チーム'),
    ('MOD-002', 'app.utils.openai_client', 'OpenAI APIとの通信を処理するユーティリティモジュール', 'openai', 'AI開発チーム'),
    ('MOD-003', 'app.prompts.chat_prompt', 'AIプロンプトテンプレートを提供するモジュール', 'なし', 'AI開発チーム'),
    ('MOD-004', 'app.schemas.llm', 'AIリクエスト/レスポンスのPydanticスキーマを定義するモジュール', 'pydantic', 'AI開発チーム'),
    ('MOD-005', 'app.dependencies', '認証依存関係を提供するモジュール', 'app.utils.auth', 'セキュリティチーム'),
    ('MOD-006', 'app.utils.logging', 'ロギング機能を提供するユーティリティモジュール', 'logging', 'インフラチーム'),
]

for i, mod in enumerate(modules, 4):
    module[f'A{i}'] = mod[0]
    module[f'B{i}'] = mod[1]
    module[f'C{i}'] = mod[2]
    module[f'D{i}'] = mod[3]
    module[f'E{i}'] = mod[4]
    
    for col in range(5):
        cell = module[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 3]:  # Description, dependencies columns
            cell.alignment = Alignment(wrap_text=True)

module.column_dimensions['A'].width = 10
module.column_dimensions['B'].width = 20
module.column_dimensions['C'].width = 40
module.column_dimensions['D'].width = 30
module.column_dimensions['E'].width = 15

file_list = wb['ファイル一覧']
file_list['A1'] = 'OpenAI連携機能 ファイル一覧'
file_list['A1'].font = title_font

file_list['A3'] = 'ID'
file_list['B3'] = 'ファイルパス'
file_list['C3'] = '説明'
file_list['D3'] = '依存ファイル'
file_list['E3'] = '行数'

for cell in file_list[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

files = [
    ('FILE-001', 'app/routes/llm.py', 'OpenAI APIエンドポイントを提供するルーターファイル', 'app/utils/openai_client.py, app/prompts/chat_prompt.py', '~50'),
    ('FILE-002', 'app/utils/openai_client.py', 'OpenAI APIとの通信を処理するユーティリティファイル', 'なし', '~70'),
    ('FILE-003', 'app/prompts/chat_prompt.py', 'AIプロンプトテンプレートを提供するファイル', 'なし', '~30'),
    ('FILE-004', 'app/schemas/llm.py', 'AIリクエスト/レスポンスのPydanticスキーマを定義するファイル', 'なし', '~40'),
    ('FILE-005', 'app/main.py', 'FastAPIアプリケーションのメインファイル（LLMルーターの登録を含む）', 'app/routes/llm.py', '~100'),
    ('FILE-006', 'requirements.txt', 'プロジェクトの依存関係を定義するファイル（openaiパッケージを含む）', 'なし', '~10'),
]

for i, file in enumerate(files, 4):
    file_list[f'A{i}'] = file[0]
    file_list[f'B{i}'] = file[1]
    file_list[f'C{i}'] = file[2]
    file_list[f'D{i}'] = file[3]
    file_list[f'E{i}'] = file[4]
    
    for col in range(5):
        cell = file_list[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 3]:  # Description, dependencies columns
            cell.alignment = Alignment(wrap_text=True)

file_list.column_dimensions['A'].width = 10
file_list.column_dimensions['B'].width = 30
file_list.column_dimensions['C'].width = 40
file_list.column_dimensions['D'].width = 30
file_list.column_dimensions['E'].width = 10

function = wb['関数仕様']
function['A1'] = 'OpenAI連携機能 関数仕様'
function['A1'].font = title_font

function['A3'] = 'ID'
function['B3'] = '関数名'
function['C3'] = '説明'
function['D3'] = '引数'
function['E3'] = '戻り値'
function['F3'] = '例外'
function['G3'] = '複雑度'

for cell in function[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

functions = [
    ('FUNC-001', 'generate_response', 
     'OpenAI APIを使用してAIレスポンスを生成する', 
     'messages: List[Dict[str, str]]\nmax_tokens: Optional[int] = 1000\ntemperature: Optional[float] = 0.7', 
     'Dict[str, Any] - レスポンステキスト、モデル、使用量情報を含む辞書', 
     'ValueError - APIキーが設定されていない場合\nException - API呼び出し中にエラーが発生した場合',
     'O(1)'),
    ('FUNC-002', 'get_chat_prompt', 
     'OpenAI API用のチャットプロンプトを生成する', 
     'user_input: str\nsystem_prompt: str = SYSTEM_PROMPT', 
     'List[Dict[str, str]] - OpenAI API用のメッセージ辞書のリスト', 
     'なし',
     'O(1)'),
    ('FUNC-003', 'chat', 
     'AIチャットエンドポイントのハンドラ関数', 
     'request: LLMRequest\ncurrent_user: User = Depends(get_current_active_user)', 
     'LLMResponse - AIからの応答を含むレスポンスオブジェクト', 
     'HTTPException - OpenAI API呼び出し中にエラーが発生した場合',
     'O(1)'),
]

for i, func in enumerate(functions, 4):
    function[f'A{i}'] = func[0]
    function[f'B{i}'] = func[1]
    function[f'C{i}'] = func[2]
    function[f'D{i}'] = func[3]
    function[f'E{i}'] = func[4]
    function[f'F{i}'] = func[5]
    function[f'G{i}'] = func[6]
    
    for col in range(7):
        cell = function[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 3, 4, 5]:  # Description, args, return, exceptions columns
            cell.alignment = Alignment(wrap_text=True)

function.column_dimensions['A'].width = 10
function.column_dimensions['B'].width = 20
function.column_dimensions['C'].width = 30
function.column_dimensions['D'].width = 30
function.column_dimensions['E'].width = 30
function.column_dimensions['F'].width = 30
function.column_dimensions['G'].width = 10

class_impl = wb['クラス実装']
class_impl['A1'] = 'OpenAI連携機能 クラス実装'
class_impl['A1'].font = title_font

class_impl['A3'] = 'ID'
class_impl['B3'] = 'クラス名'
class_impl['C3'] = '説明'
class_impl['D3'] = '属性'
class_impl['E3'] = 'メソッド'
class_impl['F3'] = '実装詳細'

for cell in class_impl[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

classes = [
    ('CLS-001', 'LLMRequest', 
     'AIリクエストのPydanticモデル', 
     'prompt: str - ユーザーの入力または質問\nmax_tokens: Optional[int] - 生成する最大トークン数\ntemperature: Optional[float] - 応答の多様性を制御するパラメータ', 
     'なし（Pydanticモデル）', 
     'Pydanticの標準的なモデル実装。Field()を使用して各フィールドに説明を追加し、Swagger UIでのドキュメント表示を改善。'),
    ('CLS-002', 'LLMResponse', 
     'AIレスポンスのPydanticモデル', 
     'response: str - AIから生成された応答テキスト\nmodel: Optional[str] - 使用されたAIモデル\nusage: Optional[Dict[str, Any]] - トークン使用量情報', 
     'なし（Pydanticモデル）', 
     'Pydanticの標準的なモデル実装。Field()を使用して各フィールドに説明を追加し、Swagger UIでのドキュメント表示を改善。'),
]

for i, cls in enumerate(classes, 4):
    class_impl[f'A{i}'] = cls[0]
    class_impl[f'B{i}'] = cls[1]
    class_impl[f'C{i}'] = cls[2]
    class_impl[f'D{i}'] = cls[3]
    class_impl[f'E{i}'] = cls[4]
    class_impl[f'F{i}'] = cls[5]
    
    for col in range(6):
        cell = class_impl[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 3, 4, 5]:  # Description, attributes, methods, implementation columns
            cell.alignment = Alignment(wrap_text=True)

class_impl.column_dimensions['A'].width = 10
class_impl.column_dimensions['B'].width = 15
class_impl.column_dimensions['C'].width = 25
class_impl.column_dimensions['D'].width = 35
class_impl.column_dimensions['E'].width = 25
class_impl.column_dimensions['F'].width = 40

database = wb['データベース']
database['A1'] = 'OpenAI連携機能 データベース設計'
database['A1'].font = title_font

database['A3'] = 'ID'
database['B3'] = 'テーブル名'
database['C3'] = '説明'
database['D3'] = 'カラム'
database['E3'] = 'データ型'
database['F3'] = '制約'

for cell in database[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

tables = [
    ('DB-001', 'なし', 
     'OpenAI連携機能は直接データベースを使用しません。ユーザー認証のためのUserテーブルは認証モジュールで定義されています。', 
     'N/A', 
     'N/A', 
     'N/A'),
]

for i, table in enumerate(tables, 4):
    database[f'A{i}'] = table[0]
    database[f'B{i}'] = table[1]
    database[f'C{i}'] = table[2]
    database[f'D{i}'] = table[3]
    database[f'E{i}'] = table[4]
    database[f'F{i}'] = table[5]
    
    for col in range(6):
        cell = database[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col == 2:  # Description column
            cell.alignment = Alignment(wrap_text=True)

database.column_dimensions['A'].width = 10
database.column_dimensions['B'].width = 15
database.column_dimensions['C'].width = 50
database.column_dimensions['D'].width = 20
database.column_dimensions['E'].width = 15
database.column_dimensions['F'].width = 20

error = wb['エラー処理']
error['A1'] = 'OpenAI連携機能 エラー処理'
error['A1'].font = title_font

error['A3'] = 'ID'
error['B3'] = 'エラーコード'
error['C3'] = '説明'
error['D3'] = '発生条件'
error['E3'] = '処理方法'
error['F3'] = 'ユーザーへの表示'

for cell in error[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

errors = [
    ('ERR-001', '401 Unauthorized', 
     '認証エラー', 
     'ユーザーが認証されていない、またはトークンが無効な場合', 
     'HTTPExceptionを発生させ、WWW-Authenticateヘッダーを設定', 
     '{"detail": "Could not validate credentials"}'),
    ('ERR-002', '422 Unprocessable Entity', 
     'リクエスト検証エラー', 
     'リクエストボディがLLMRequestスキーマに適合しない場合', 
     'FastAPIの自動検証機能によりHTTPExceptionが発生', 
     '{"detail": [{"loc": ["body", "prompt"], "msg": "field required", "type": "value_error.missing"}]}'),
    ('ERR-003', '500 Internal Server Error', 
     'OpenAI API呼び出しエラー', 
     'APIキーが設定されていない、またはAPI呼び出し中にエラーが発生した場合', 
     'try-except文でエラーをキャッチし、ログに記録してHTTPExceptionを発生', 
     '{"detail": "Error generating OpenAI response"}'),
    ('ERR-004', '429 Too Many Requests', 
     'レート制限エラー', 
     'OpenAI APIのレート制限に達した場合', 
     'try-except文でエラーをキャッチし、ログに記録してHTTPExceptionを発生', 
     '{"detail": "OpenAI API rate limit exceeded. Please try again later."}'),
]

for i, err in enumerate(errors, 4):
    error[f'A{i}'] = err[0]
    error[f'B{i}'] = err[1]
    error[f'C{i}'] = err[2]
    error[f'D{i}'] = err[3]
    error[f'E{i}'] = err[4]
    error[f'F{i}'] = err[5]
    
    for col in range(6):
        cell = error[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [3, 4, 5]:  # Condition, handling, display columns
            cell.alignment = Alignment(wrap_text=True)

error.column_dimensions['A'].width = 10
error.column_dimensions['B'].width = 20
error.column_dimensions['C'].width = 20
error.column_dimensions['D'].width = 30
error.column_dimensions['E'].width = 30
error.column_dimensions['F'].width = 40

test = wb['テスト仕様']
test['A1'] = 'OpenAI連携機能 テスト仕様'
test['A1'].font = title_font

test['A3'] = 'ID'
test['B3'] = 'テスト名'
test['C3'] = '説明'
test['D3'] = '前提条件'
test['E3'] = 'テスト手順'
test['F3'] = '期待結果'
test['G3'] = 'テストタイプ'

for cell in test[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

tests = [
    ('TEST-001', 'AI問い合わせ正常系テスト', 
     'ユーザーが認証された状態でAI問い合わせを行い、正常にレスポンスが返ることを確認', 
     '- ユーザーが登録済み\n- ユーザーが認証済み\n- OpenAI APIキーが設定済み', 
     '1. /auth/tokenエンドポイントでトークンを取得\n2. 取得したトークンをAuthorizationヘッダーに設定\n3. /llm/chatエンドポイントにリクエストを送信\n4. レスポンスを確認', 
     '- ステータスコード: 200 OK\n- レスポンスボディ: LLMResponseスキーマに適合\n- response, model, usageフィールドが存在',
     '単体テスト'),
    ('TEST-002', '認証エラーテスト', 
     '認証されていないユーザーがAI問い合わせを行った場合にエラーが返ることを確認', 
     '- OpenAI APIキーが設定済み', 
     '1. 認証トークンなしで/llm/chatエンドポイントにリクエストを送信\n2. レスポンスを確認', 
     '- ステータスコード: 401 Unauthorized\n- レスポンスボディ: {"detail": "Not authenticated"}',
     '単体テスト'),
    ('TEST-003', 'リクエスト検証エラーテスト', 
     '不正なリクエストボディでAI問い合わせを行った場合にエラーが返ることを確認', 
     '- ユーザーが登録済み\n- ユーザーが認証済み\n- OpenAI APIキーが設定済み', 
     '1. /auth/tokenエンドポイントでトークンを取得\n2. 取得したトークンをAuthorizationヘッダーに設定\n3. promptフィールドを含まない不正なリクエストボディで/llm/chatエンドポイントにリクエストを送信\n4. レスポンスを確認', 
     '- ステータスコード: 422 Unprocessable Entity\n- レスポンスボディに検証エラーの詳細が含まれる',
     '単体テスト'),
    ('TEST-004', 'OpenAI APIキー未設定テスト', 
     'OpenAI APIキーが設定されていない場合にエラーが返ることを確認', 
     '- ユーザーが登録済み\n- ユーザーが認証済み\n- OpenAI APIキーが未設定', 
     '1. /auth/tokenエンドポイントでトークンを取得\n2. 取得したトークンをAuthorizationヘッダーに設定\n3. /llm/chatエンドポイントにリクエストを送信\n4. レスポンスを確認', 
     '- ステータスコード: 500 Internal Server Error\n- レスポンスボディ: {"detail": "Error generating OpenAI response"}',
     '単体テスト'),
    ('TEST-005', 'パフォーマンステスト', 
     'AI問い合わせの応答時間を測定', 
     '- ユーザーが登録済み\n- ユーザーが認証済み\n- OpenAI APIキーが設定済み', 
     '1. /auth/tokenエンドポイントでトークンを取得\n2. 取得したトークンをAuthorizationヘッダーに設定\n3. /llm/chatエンドポイントに10回連続でリクエストを送信\n4. 各リクエストの応答時間を測定', 
     '- 平均応答時間が5秒以内\n- 最大応答時間が10秒以内',
     '性能テスト'),
    ('TEST-006', '負荷テスト', 
     '複数の同時リクエストを処理できることを確認', 
     '- ユーザーが登録済み\n- ユーザーが認証済み\n- OpenAI APIキーが設定済み', 
     '1. /auth/tokenエンドポイントでトークンを取得\n2. 取得したトークンをAuthorizationヘッダーに設定\n3. 10の並列スレッドから/llm/chatエンドポイントに同時にリクエストを送信\n4. すべてのレスポンスを確認', 
     '- すべてのリクエストが成功（ステータスコード: 200 OK）\n- エラーが発生しない',
     '負荷テスト'),
]

for i, test_case in enumerate(tests, 4):
    test[f'A{i}'] = test_case[0]
    test[f'B{i}'] = test_case[1]
    test[f'C{i}'] = test_case[2]
    test[f'D{i}'] = test_case[3]
    test[f'E{i}'] = test_case[4]
    test[f'F{i}'] = test_case[5]
    test[f'G{i}'] = test_case[6]
    
    for col in range(7):
        cell = test[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 3, 4, 5]:  # Description, preconditions, steps, expected columns
            cell.alignment = Alignment(wrap_text=True)

test.column_dimensions['A'].width = 10
test.column_dimensions['B'].width = 25
test.column_dimensions['C'].width = 30
test.column_dimensions['D'].width = 30
test.column_dimensions['E'].width = 40
test.column_dimensions['F'].width = 40
test.column_dimensions['G'].width = 15

os.makedirs('DOC/単体テスト', exist_ok=True)

wb.save('DOC/単体テスト/OpenAI連携機能_プログラム詳細設計書.xlsx')
print('Excel file created successfully: DOC/単体テスト/OpenAI連携機能_プログラム詳細設計書.xlsx')
