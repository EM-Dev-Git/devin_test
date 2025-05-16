import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

wb = openpyxl.Workbook()

sheets = [
    '表紙', '改訂履歴', 'API仕様', 'データフロー', 
    'シーケンス図', 'エラー処理', 'UI設計'
]

wb.remove(wb.active)
for sheet_name in sheets:
    wb.create_sheet(sheet_name)

title_font = Font(name='Yu Gothic', size=16, bold=True)
header_font = Font(name='Yu Gothic', size=12, bold=True)
normal_font = Font(name='Yu Gothic', size=11)
header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

cover = wb['表紙']
cover['B2'] = 'OpenAI連携機能 外部設計書'
cover['B2'].font = Font(name='Yu Gothic', size=20, bold=True)
cover['B4'] = 'プロジェクト名：'
cover['C4'] = 'EM_test_project'
cover['B5'] = '作成日：'
cover['C5'] = datetime.now().strftime('%Y年%m月%d日')
cover['B6'] = '作成者：'
cover['C6'] = 'Devin AI'
cover['B8'] = '概要：'
cover['C8'] = 'このドキュメントはOpenAI APIを使用したAI問い合わせ機能の外部設計を定義します。'
cover['C9'] = 'API仕様、データフロー、シーケンス図、エラー処理、UI設計を含みます。'

for col in range(1, 10):
    cover.column_dimensions[get_column_letter(col)].width = 15

history = wb['改訂履歴']
history['A1'] = '改訂履歴'
history['A1'].font = title_font
history['A3'] = 'バージョン'
history['B3'] = '日付'
history['C3'] = '作成者'
history['D3'] = '変更内容'
history['E3'] = '承認者'

for cell in history[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

history['A4'] = '1.0'
history['B4'] = datetime.now().strftime('%Y/%m/%d')
history['C4'] = 'Devin AI'
history['D4'] = '初版作成'
history['E4'] = ''

for cell in history[4]:
    cell.font = normal_font
    cell.border = border

history.column_dimensions['A'].width = 12
history.column_dimensions['B'].width = 12
history.column_dimensions['C'].width = 15
history.column_dimensions['D'].width = 40
history.column_dimensions['E'].width = 15

api = wb['API仕様']
api['A1'] = 'OpenAI連携機能 API仕様'
api['A1'].font = title_font

api['A3'] = 'エンドポイント'
api['B3'] = 'メソッド'
api['C3'] = '説明'
api['D3'] = 'リクエストパラメータ'
api['E3'] = 'レスポンス'
api['F3'] = '認証'
api['G3'] = 'ステータスコード'

for cell in api[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

endpoints = [
    ('/llm/chat', 'POST', 'AIとチャットするためのエンドポイント', 
     '{\n  "prompt": "string",\n  "max_tokens": 1000,\n  "temperature": 0.7\n}', 
     '{\n  "response": "string",\n  "model": "string",\n  "usage": {\n    "prompt_tokens": 0,\n    "completion_tokens": 0,\n    "total_tokens": 0\n  }\n}', 
     'JWT Bearer Token', 
     '200 OK, 401 Unauthorized, 422 Unprocessable Entity, 500 Internal Server Error'),
]

for i, endpoint in enumerate(endpoints, 4):
    api[f'A{i}'] = endpoint[0]
    api[f'B{i}'] = endpoint[1]
    api[f'C{i}'] = endpoint[2]
    api[f'D{i}'] = endpoint[3]
    api[f'E{i}'] = endpoint[4]
    api[f'F{i}'] = endpoint[5]
    api[f'G{i}'] = endpoint[6]
    
    for col in range(7):
        cell = api[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 3, 4, 6]:  # Description, request, response, status code columns
            cell.alignment = Alignment(wrap_text=True)

api.column_dimensions['A'].width = 15
api.column_dimensions['B'].width = 10
api.column_dimensions['C'].width = 30
api.column_dimensions['D'].width = 30
api.column_dimensions['E'].width = 30
api.column_dimensions['F'].width = 15
api.column_dimensions['G'].width = 20

data_flow = wb['データフロー']
data_flow['A1'] = 'OpenAI連携機能 データフロー'
data_flow['A1'].font = title_font

data_flow['A3'] = 'ID'
data_flow['B3'] = 'フロー名'
data_flow['C3'] = '説明'
data_flow['D3'] = '入力元'
data_flow['E3'] = '出力先'
data_flow['F3'] = 'データ形式'

for cell in data_flow[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

flows = [
    ('DF-001', 'ユーザー入力→OpenAI', 'ユーザーの入力テキストをOpenAI APIに送信', 'クライアント', 'OpenAI API', 'JSON'),
    ('DF-002', 'OpenAI→ユーザー', 'OpenAI APIからの応答をユーザーに返す', 'OpenAI API', 'クライアント', 'JSON'),
    ('DF-003', 'ユーザー認証', 'JWTトークンによるユーザー認証', 'クライアント', 'サーバー', 'JWT'),
    ('DF-004', 'ログ記録', 'リクエストとレスポンスのログ記録', 'サーバー', 'ログファイル', 'テキスト'),
]

for i, flow in enumerate(flows, 4):
    data_flow[f'A{i}'] = flow[0]
    data_flow[f'B{i}'] = flow[1]
    data_flow[f'C{i}'] = flow[2]
    data_flow[f'D{i}'] = flow[3]
    data_flow[f'E{i}'] = flow[4]
    data_flow[f'F{i}'] = flow[5]
    
    for col in range(6):
        cell = data_flow[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col == 2:  # Description column
            cell.alignment = Alignment(wrap_text=True)

data_flow.column_dimensions['A'].width = 10
data_flow.column_dimensions['B'].width = 20
data_flow.column_dimensions['C'].width = 40
data_flow.column_dimensions['D'].width = 15
data_flow.column_dimensions['E'].width = 15
data_flow.column_dimensions['F'].width = 15

sequence = wb['シーケンス図']
sequence['A1'] = 'OpenAI連携機能 シーケンス図'
sequence['A1'].font = title_font

sequence['A3'] = 'ID'
sequence['B3'] = 'シーケンス名'
sequence['C3'] = '説明'
sequence['D3'] = 'アクター'
sequence['E3'] = 'シーケンスステップ'

for cell in sequence[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

sequences = [
    ('SQ-001', 'AI問い合わせ', 'ユーザーがAIに問い合わせを行うシーケンス', 'ユーザー, サーバー, OpenAI API', 
     '1. ユーザーがログイン\n2. ユーザーがプロンプトを入力\n3. サーバーがリクエストを検証\n4. サーバーがOpenAI APIにリクエスト送信\n5. OpenAI APIがレスポンスを生成\n6. サーバーがレスポンスを受信\n7. サーバーがレスポンスをユーザーに返す\n8. サーバーがログを記録'),
    ('SQ-002', 'エラーハンドリング', 'APIエラー発生時のシーケンス', 'ユーザー, サーバー, OpenAI API', 
     '1. ユーザーがプロンプトを入力\n2. サーバーがリクエストを検証\n3. サーバーがOpenAI APIにリクエスト送信\n4. OpenAI APIがエラーを返す\n5. サーバーがエラーを処理\n6. サーバーがエラーメッセージをユーザーに返す\n7. サーバーがエラーログを記録'),
]

for i, seq in enumerate(sequences, 4):
    sequence[f'A{i}'] = seq[0]
    sequence[f'B{i}'] = seq[1]
    sequence[f'C{i}'] = seq[2]
    sequence[f'D{i}'] = seq[3]
    sequence[f'E{i}'] = seq[4]
    
    for col in range(5):
        cell = sequence[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 3, 4]:  # Description, actors, steps columns
            cell.alignment = Alignment(wrap_text=True)

sequence.column_dimensions['A'].width = 10
sequence.column_dimensions['B'].width = 20
sequence.column_dimensions['C'].width = 30
sequence.column_dimensions['D'].width = 25
sequence.column_dimensions['E'].width = 50

error = wb['エラー処理']
error['A1'] = 'OpenAI連携機能 エラー処理'
error['A1'].font = title_font

error['A3'] = 'ID'
error['B3'] = 'エラー種類'
error['C3'] = '説明'
error['D3'] = 'ステータスコード'
error['E3'] = 'エラーメッセージ'
error['F3'] = '対応方法'

for cell in error[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

errors = [
    ('ERR-001', '認証エラー', 'ユーザーが認証されていない', '401', 'Could not validate credentials', 'ログイン画面にリダイレクト'),
    ('ERR-002', 'バリデーションエラー', 'リクエストパラメータが無効', '422', 'Validation Error', 'エラーメッセージを表示し、正しい入力を促す'),
    ('ERR-003', 'OpenAI APIエラー', 'OpenAI APIからのエラーレスポンス', '500', 'Error generating OpenAI response', 'エラーメッセージを表示し、再試行を促す'),
    ('ERR-004', 'APIキーエラー', 'OpenAI APIキーが設定されていない', '500', 'OpenAI API key not configured', '管理者に連絡するよう促す'),
    ('ERR-005', 'レート制限エラー', 'OpenAI APIのレート制限に達した', '429', 'Rate limit exceeded', '一定時間後に再試行するよう促す'),
]

for i, err in enumerate(errors, 4):
    error[f'A{i}'] = err[0]
    error[f'B{i}'] = err[1]
    error[f'C{i}'] = err[2]
    error[f'D{i}'] = err[3]
    error[f'E{i}'] = err[4]
    error[f'F{i}'] = err[5]
    
    for col in range(6):
        cell = error[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 4, 5]:  # Description, message, handling columns
            cell.alignment = Alignment(wrap_text=True)

error.column_dimensions['A'].width = 10
error.column_dimensions['B'].width = 20
error.column_dimensions['C'].width = 30
error.column_dimensions['D'].width = 15
error.column_dimensions['E'].width = 30
error.column_dimensions['F'].width = 30

ui = wb['UI設計']
ui['A1'] = 'OpenAI連携機能 UI設計'
ui['A1'].font = title_font

ui['A3'] = 'ID'
ui['B3'] = '画面名'
ui['C3'] = '説明'
ui['D3'] = 'UI要素'
ui['E3'] = '入力検証'
ui['F3'] = 'アクション'

for cell in ui[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

designs = [
    ('UI-001', 'AI問い合わせ画面', 'ユーザーがAIに問い合わせを行う画面', 
     'テキスト入力フィールド, 送信ボタン, レスポンス表示エリア, 設定パネル', 
     'プロンプトは必須, 最大1000文字まで', 
     '送信ボタンクリック: プロンプトをサーバーに送信\n設定パネル: max_tokensとtemperatureを調整可能'),
    ('UI-002', 'ログイン画面', 'ユーザー認証のための画面', 
     'ユーザー名入力フィールド, パスワード入力フィールド, ログインボタン', 
     'ユーザー名とパスワードは必須', 
     'ログインボタンクリック: 認証情報をサーバーに送信'),
    ('UI-003', 'エラー表示', 'エラーメッセージを表示するモーダル', 
     'エラータイトル, エラーメッセージ, 閉じるボタン', 
     'なし', 
     '閉じるボタンクリック: モーダルを閉じる'),
]

for i, design in enumerate(designs, 4):
    ui[f'A{i}'] = design[0]
    ui[f'B{i}'] = design[1]
    ui[f'C{i}'] = design[2]
    ui[f'D{i}'] = design[3]
    ui[f'E{i}'] = design[4]
    ui[f'F{i}'] = design[5]
    
    for col in range(6):
        cell = ui[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 3, 4, 5]:  # Description, UI elements, validation, actions columns
            cell.alignment = Alignment(wrap_text=True)

ui.column_dimensions['A'].width = 10
ui.column_dimensions['B'].width = 20
ui.column_dimensions['C'].width = 30
ui.column_dimensions['D'].width = 30
ui.column_dimensions['E'].width = 25
ui.column_dimensions['F'].width = 30

os.makedirs('DOC/外部設計', exist_ok=True)

wb.save('DOC/外部設計/OpenAI連携機能_外部設計書.xlsx')
print('Excel file created successfully: DOC/外部設計/OpenAI連携機能_外部設計書.xlsx')
