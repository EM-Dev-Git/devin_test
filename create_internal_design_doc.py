import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

wb = openpyxl.Workbook()

sheets = [
    '表紙', '改訂履歴', 'システム構成', 'コンポーネント設計', 
    'クラス設計', 'シーケンス詳細', 'データ構造', 'アルゴリズム'
]

wb.remove(wb.active)
for sheet_name in sheets:
    wb.create_sheet(sheet_name)

title_font = Font(name='Yu Gothic', size=16, bold=True)
header_font = Font(name='Yu Gothic', size=12, bold=True)
normal_font = Font(name='Yu Gothic', size=11)
header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

cover = wb['表紙']
cover['B2'] = 'OpenAI連携機能 内部設計書'
cover['B2'].font = Font(name='Yu Gothic', size=20, bold=True)
cover['B4'] = 'プロジェクト名：'
cover['C4'] = 'EM_test_project'
cover['B5'] = '作成日：'
cover['C5'] = datetime.now().strftime('%Y年%m月%d日')
cover['B6'] = '作成者：'
cover['C6'] = 'Devin AI'
cover['B8'] = '概要：'
cover['C8'] = 'このドキュメントはOpenAI APIを使用したAI問い合わせ機能の内部設計を定義します。'
cover['C9'] = 'システム構成、コンポーネント設計、クラス設計、データ構造、アルゴリズムを含みます。'

for col in range(1, 10):
    cover.column_dimensions[get_column_letter(col)].width = 15

history = wb['改訂履歴']
history['A1'] = '改訂履歴'
history['A1'].font = title_font
history['A3'] = 'バージョン'
history['B3'] = '日付'
history['C3'] = '作成者'
history['D3'] = '変更内容'
history['E3'] = '承認者'

for cell in history[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

history['A4'] = '1.0'
history['B4'] = datetime.now().strftime('%Y/%m/%d')
history['C4'] = 'Devin AI'
history['D4'] = '初版作成'
history['E4'] = ''

for cell in history[4]:
    cell.font = normal_font
    cell.border = border

history.column_dimensions['A'].width = 12
history.column_dimensions['B'].width = 12
history.column_dimensions['C'].width = 15
history.column_dimensions['D'].width = 40
history.column_dimensions['E'].width = 15

architecture = wb['システム構成']
architecture['A1'] = 'OpenAI連携機能 システム構成'
architecture['A1'].font = title_font

architecture['A3'] = 'ID'
architecture['B3'] = 'コンポーネント名'
architecture['C3'] = '説明'
architecture['D3'] = '依存関係'
architecture['E3'] = '技術スタック'

for cell in architecture[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

components = [
    ('COMP-001', 'FastAPIアプリケーション', 'メインのWebアプリケーションフレームワーク', 'なし', 'FastAPI, Uvicorn'),
    ('COMP-002', 'OpenAIクライアント', 'OpenAI APIとの通信を処理するモジュール', 'COMP-001', 'OpenAI Python SDK'),
    ('COMP-003', 'プロンプトテンプレート', 'AIへの入力に使用するテンプレートを管理', 'COMP-002', 'Python文字列テンプレート'),
    ('COMP-004', '認証モジュール', 'ユーザー認証とトークン管理', 'COMP-001', 'JWT, Passlib'),
    ('COMP-005', 'ロギングモジュール', 'アプリケーションログの管理', 'COMP-001', 'Python標準ロギング'),
    ('COMP-006', 'データベース', 'ユーザー情報の永続化', 'COMP-001, COMP-004', 'SQLite, SQLAlchemy'),
]

for i, comp in enumerate(components, 4):
    architecture[f'A{i}'] = comp[0]
    architecture[f'B{i}'] = comp[1]
    architecture[f'C{i}'] = comp[2]
    architecture[f'D{i}'] = comp[3]
    architecture[f'E{i}'] = comp[4]
    
    for col in range(5):
        cell = architecture[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col == 2:  # Description column
            cell.alignment = Alignment(wrap_text=True)

architecture.column_dimensions['A'].width = 10
architecture.column_dimensions['B'].width = 25
architecture.column_dimensions['C'].width = 40
architecture.column_dimensions['D'].width = 20
architecture.column_dimensions['E'].width = 20

component = wb['コンポーネント設計']
component['A1'] = 'OpenAI連携機能 コンポーネント設計'
component['A1'].font = title_font

component['A3'] = 'ID'
component['B3'] = 'コンポーネント名'
component['C3'] = '責務'
component['D3'] = 'インターフェース'
component['E3'] = '内部構造'
component['F3'] = '注意点'

for cell in component[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

designs = [
    ('CD-001', 'OpenAIクライアント', 
     'OpenAI APIとの通信を処理し、AIレスポンスを生成する', 
     'generate_response(messages, max_tokens, temperature)', 
     'OpenAI APIクライアントの初期化、エラーハンドリング、レスポンス処理', 
     'APIキーは環境変数から取得し、コード内に直接記述しない'),
    ('CD-002', 'プロンプトテンプレート', 
     'AIへの入力に使用するテンプレートを管理する', 
     'get_chat_prompt(user_input, system_prompt)', 
     'システムプロンプトとユーザープロンプトの組み合わせ', 
     'プロンプトは別ファイルで管理し、容易に変更できるようにする'),
    ('CD-003', 'LLMルーター', 
     'AIチャットエンドポイントを提供する', 
     'POST /llm/chat', 
     'リクエスト検証、OpenAIクライアント呼び出し、レスポンス生成', 
     'すべてのエンドポイントは認証が必要'),
]

for i, design in enumerate(designs, 4):
    component[f'A{i}'] = design[0]
    component[f'B{i}'] = design[1]
    component[f'C{i}'] = design[2]
    component[f'D{i}'] = design[3]
    component[f'E{i}'] = design[4]
    component[f'F{i}'] = design[5]
    
    for col in range(6):
        cell = component[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 3, 4, 5]:  # Responsibility, interface, structure, notes columns
            cell.alignment = Alignment(wrap_text=True)

component.column_dimensions['A'].width = 10
component.column_dimensions['B'].width = 20
component.column_dimensions['C'].width = 30
component.column_dimensions['D'].width = 30
component.column_dimensions['E'].width = 30
component.column_dimensions['F'].width = 30

class_design = wb['クラス設計']
class_design['A1'] = 'OpenAI連携機能 クラス設計'
class_design['A1'].font = title_font

class_design['A3'] = 'ID'
class_design['B3'] = 'クラス名'
class_design['C3'] = '説明'
class_design['D3'] = '属性'
class_design['E3'] = 'メソッド'
class_design['F3'] = '関連クラス'

for cell in class_design[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

classes = [
    ('CLS-001', 'LLMRequest', 
     'AIリクエストのデータモデル', 
     'prompt: str\nmax_tokens: Optional[int]\ntemperature: Optional[float]', 
     'なし（Pydanticモデル）', 
     'なし'),
    ('CLS-002', 'LLMResponse', 
     'AIレスポンスのデータモデル', 
     'response: str\nmodel: Optional[str]\nusage: Optional[Dict[str, Any]]', 
     'なし（Pydanticモデル）', 
     'なし'),
    ('CLS-003', 'OpenAIClient', 
     'OpenAI APIとの通信を処理するクラス', 
     'client: OpenAI', 
     'generate_response(messages, max_tokens, temperature)', 
     'LLMRequest, LLMResponse'),
]

for i, cls in enumerate(classes, 4):
    class_design[f'A{i}'] = cls[0]
    class_design[f'B{i}'] = cls[1]
    class_design[f'C{i}'] = cls[2]
    class_design[f'D{i}'] = cls[3]
    class_design[f'E{i}'] = cls[4]
    class_design[f'F{i}'] = cls[5]
    
    for col in range(6):
        cell = class_design[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 3, 4]:  # Description, attributes, methods columns
            cell.alignment = Alignment(wrap_text=True)

class_design.column_dimensions['A'].width = 10
class_design.column_dimensions['B'].width = 20
class_design.column_dimensions['C'].width = 30
class_design.column_dimensions['D'].width = 30
class_design.column_dimensions['E'].width = 30
class_design.column_dimensions['F'].width = 20

sequence = wb['シーケンス詳細']
sequence['A1'] = 'OpenAI連携機能 シーケンス詳細'
sequence['A1'].font = title_font

sequence['A3'] = 'ID'
sequence['B3'] = 'シーケンス名'
sequence['C3'] = '説明'
sequence['D3'] = '詳細ステップ'
sequence['E3'] = '例外フロー'

for cell in sequence[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

sequences = [
    ('SQD-001', 'AI問い合わせ処理', 'ユーザーからのAI問い合わせ処理の詳細シーケンス', 
     '1. LLMルーターがリクエストを受信\n2. リクエストをLLMRequestモデルで検証\n3. get_current_active_user()でユーザー認証\n4. get_chat_prompt()でプロンプトテンプレート生成\n5. generate_response()でOpenAI API呼び出し\n6. レスポンスをLLMResponseモデルに変換\n7. JSONレスポンスを返す\n8. ログを記録', 
     '- ユーザー認証失敗: 401エラー返却\n- リクエスト検証失敗: 422エラー返却\n- OpenAI API呼び出し失敗: 500エラー返却、エラーログ記録'),
    ('SQD-002', 'OpenAI API呼び出し', 'OpenAI APIへのリクエスト送信と応答処理の詳細', 
     '1. APIキーの存在確認\n2. メッセージ配列の構築\n3. OpenAI ChatCompletionリクエスト作成\n4. APIリクエスト送信\n5. レスポンス受信\n6. レスポンスデータの抽出\n7. 構造化されたレスポンスの返却', 
     '- APIキー未設定: ValueErrorを発生\n- API接続エラー: 例外をキャッチしてログ記録、再スロー\n- レート制限エラー: 429エラーをキャッチし、適切なエラーメッセージを返却'),
]

for i, seq in enumerate(sequences, 4):
    sequence[f'A{i}'] = seq[0]
    sequence[f'B{i}'] = seq[1]
    sequence[f'C{i}'] = seq[2]
    sequence[f'D{i}'] = seq[3]
    sequence[f'E{i}'] = seq[4]
    
    for col in range(5):
        cell = sequence[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 3, 4]:  # Description, steps, exceptions columns
            cell.alignment = Alignment(wrap_text=True)

sequence.column_dimensions['A'].width = 10
sequence.column_dimensions['B'].width = 20
sequence.column_dimensions['C'].width = 30
sequence.column_dimensions['D'].width = 50
sequence.column_dimensions['E'].width = 40

data = wb['データ構造']
data['A1'] = 'OpenAI連携機能 データ構造'
data['A1'].font = title_font

data['A3'] = 'ID'
data['B3'] = 'データ構造名'
data['C3'] = '説明'
data['D3'] = 'フィールド'
data['E3'] = 'データ型'
data['F3'] = '制約'

for cell in data[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

structures = [
    ('DS-001', 'LLMRequest', 'AIリクエストのデータ構造', 'prompt', 'string', '必須'),
    ('DS-001', '', '', 'max_tokens', 'integer', 'オプション、デフォルト: 1000'),
    ('DS-001', '', '', 'temperature', 'float', 'オプション、デフォルト: 0.7、範囲: 0.0-1.0'),
    ('DS-002', 'LLMResponse', 'AIレスポンスのデータ構造', 'response', 'string', '必須'),
    ('DS-002', '', '', 'model', 'string', 'オプション'),
    ('DS-002', '', '', 'usage', 'object', 'オプション'),
    ('DS-002', '', '', 'usage.prompt_tokens', 'integer', 'オプション'),
    ('DS-002', '', '', 'usage.completion_tokens', 'integer', 'オプション'),
    ('DS-002', '', '', 'usage.total_tokens', 'integer', 'オプション'),
    ('DS-003', 'ChatPrompt', 'OpenAI APIに送信するメッセージ構造', 'role', 'string', '必須、"system"または"user"'),
    ('DS-003', '', '', 'content', 'string', '必須'),
]

for i, struct in enumerate(structures, 4):
    data[f'A{i}'] = struct[0]
    data[f'B{i}'] = struct[1]
    data[f'C{i}'] = struct[2]
    data[f'D{i}'] = struct[3]
    data[f'E{i}'] = struct[4]
    data[f'F{i}'] = struct[5]
    
    for col in range(6):
        cell = data[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 5]:  # Description, constraints columns
            cell.alignment = Alignment(wrap_text=True)

data.column_dimensions['A'].width = 10
data.column_dimensions['B'].width = 15
data.column_dimensions['C'].width = 30
data.column_dimensions['D'].width = 20
data.column_dimensions['E'].width = 15
data.column_dimensions['F'].width = 30

algorithm = wb['アルゴリズム']
algorithm['A1'] = 'OpenAI連携機能 アルゴリズム'
algorithm['A1'].font = title_font

algorithm['A3'] = 'ID'
algorithm['B3'] = 'アルゴリズム名'
algorithm['C3'] = '説明'
algorithm['D3'] = '擬似コード'
algorithm['E3'] = '複雑度'
algorithm['F3'] = '注意点'

for cell in algorithm[3]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

algorithms = [
    ('ALG-001', 'プロンプト生成アルゴリズム', 
     'ユーザー入力からOpenAI APIに送信するプロンプトを生成する', 
     'function get_chat_prompt(user_input, system_prompt):\n  messages = []\n  messages.append({"role": "system", "content": system_prompt})\n  messages.append({"role": "user", "content": format_user_prompt(user_input)})\n  return messages', 
     'O(1)', 
     'システムプロンプトは固定だが、将来的に動的に変更できるようにする'),
    ('ALG-002', 'OpenAI API呼び出しアルゴリズム', 
     'OpenAI APIを呼び出してAIレスポンスを生成する', 
     'async function generate_response(messages, max_tokens, temperature):\n  validate_api_key()\n  try:\n    response = await client.chat.completions.create(\n      model="gpt-3.5-turbo",\n      messages=messages,\n      max_tokens=max_tokens,\n      temperature=temperature\n    )\n    return format_response(response)\n  catch (error):\n    log_error(error)\n    throw error', 
     'O(1) (APIレスポンス時間に依存)', 
     'エラーハンドリングを適切に行い、APIキーの存在確認を必ず行う'),
    ('ALG-003', 'エラーハンドリングアルゴリズム', 
     'OpenAI API呼び出し中のエラーを処理する', 
     'function handle_openai_error(error):\n  if error.type == "api_key_error":\n    return {"error": "API key not configured", "status_code": 500}\n  else if error.type == "rate_limit_error":\n    return {"error": "Rate limit exceeded", "status_code": 429}\n  else:\n    log_detailed_error(error)\n    return {"error": "Error generating OpenAI response", "status_code": 500}', 
     'O(1)', 
     'エラータイプに応じて適切なステータスコードとメッセージを返す'),
]

for i, alg in enumerate(algorithms, 4):
    algorithm[f'A{i}'] = alg[0]
    algorithm[f'B{i}'] = alg[1]
    algorithm[f'C{i}'] = alg[2]
    algorithm[f'D{i}'] = alg[3]
    algorithm[f'E{i}'] = alg[4]
    algorithm[f'F{i}'] = alg[5]
    
    for col in range(6):
        cell = algorithm[f'{chr(65+col)}{i}']
        cell.font = normal_font
        cell.border = border
        if col in [2, 3, 5]:  # Description, pseudocode, notes columns
            cell.alignment = Alignment(wrap_text=True)

algorithm.column_dimensions['A'].width = 10
algorithm.column_dimensions['B'].width = 25
algorithm.column_dimensions['C'].width = 30
algorithm.column_dimensions['D'].width = 50
algorithm.column_dimensions['E'].width = 10
algorithm.column_dimensions['F'].width = 30

os.makedirs('DOC/内部設定', exist_ok=True)

wb.save('DOC/内部設定/OpenAI連携機能_内部設計書.xlsx')
print('Excel file created successfully: DOC/内部設定/OpenAI連携機能_内部設計書.xlsx')
