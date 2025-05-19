import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_unit_test_spec():
    """単体テスト仕様書を作成する"""
    
    wb = openpyxl.Workbook()
    
    ws = wb.active
    ws.title = "単体テスト仕様"
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 35
    
    header_font = Font(name='メイリオ', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type='solid', fgColor="4472C4")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    normal_font = Font(name='メイリオ', size=11)
    normal_alignment = Alignment(vertical='top', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["テストID", "テスト対象", "テスト条件", "テスト手順", "期待結果", "優先度", "備考"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    test_cases = [
        ('AUTH-001', 'ユーザー登録', 
         '有効なユーザー名、メールアドレス、パスワードを指定', 
         '1. /api/v1/auth/registerエンドポイントにPOSTリクエストを送信\n2. 有効なユーザー情報（username, email, password）を指定', 
         '1. ステータスコード201が返される\n2. レスポンスボディに成功メッセージが含まれる\n3. データベースに新しいユーザーが作成される', 
         '高', 
         'ユーザーの基本的な登録機能の確認。セキュリティ上重要なため優先度高。'),
        
        ('AUTH-002', 'ユーザー登録（重複）', 
         '既に存在するユーザー名またはメールアドレスを指定', 
         '1. 同じユーザー名またはメールアドレスで2回登録を試みる', 
         '1. 2回目の登録でステータスコード400が返される\n2. レスポンスにエラーメッセージが含まれる', 
         '高', 
         'ユーザー情報の一意性を確保するため重要。'),
        
        ('AUTH-003', 'ログイン（トークン取得）', 
         '有効なユーザー名とパスワードを指定', 
         '1. /api/v1/auth/tokenエンドポイントにPOSTリクエストを送信\n2. 有効なユーザー認証情報を指定', 
         '1. ステータスコード200が返される\n2. レスポンスにアクセストークンが含まれる\n3. トークンの形式が有効なJWTである', 
         '高', 
         'すべての認証済みエンドポイントの前提となるため優先度高。'),
        
        ('AUTH-004', 'ログイン（無効な認証情報）', 
         '無効なユーザー名またはパスワードを指定', 
         '1. /api/v1/auth/tokenエンドポイントにPOSTリクエストを送信\n2. 無効なユーザー名またはパスワードを指定', 
         '1. ステータスコード401が返される\n2. レスポンスにエラーメッセージが含まれる\n3. WWW-Authenticateヘッダーが設定される', 
         '高', 
         'セキュリティ上重要なため優先度高。'),
        
        ('ITEM-001', 'アイテム作成', 
         '認証済みユーザー、有効なアイテムデータ', 
         '1. 有効なJWTトークンをAuthorizationヘッダーに設定\n2. /api/v1/items/エンドポイントにPOSTリクエストを送信\n3. 有効なアイテムデータを指定', 
         '1. ステータスコード201が返される\n2. レスポンスにアイテムIDが含まれる\n3. レスポンスに指定したデータが含まれる\n4. データベースに新しいアイテムが作成される', 
         '高', 
         'アイテム管理の基本機能として重要。'),
        
        ('ITEM-002', 'アイテム一覧取得', 
         '認証済みユーザー、複数のアイテムが存在', 
         '1. 有効なJWTトークンをAuthorizationヘッダーに設定\n2. /api/v1/items/エンドポイントにGETリクエストを送信', 
         '1. ステータスコード200が返される\n2. レスポンスが配列形式である\n3. 現在のユーザーが所有するすべてのアイテムが含まれる', 
         '中', 
         'ページネーションの動作も確認する。'),
        
        ('ITEM-003', '特定アイテム取得', 
         '認証済みユーザー、存在するアイテムID', 
         '1. 有効なJWTトークンをAuthorizationヘッダーに設定\n2. /api/v1/items/{item_id}エンドポイントにGETリクエストを送信', 
         '1. ステータスコード200が返される\n2. レスポンスに指定したアイテムのデータが含まれる', 
         '中', 
         'アイテムの詳細表示機能の確認。'),
        
        ('ITEM-004', '存在しないアイテム取得', 
         '認証済みユーザー、存在しないアイテムID', 
         '1. 有効なJWTトークンをAuthorizationヘッダーに設定\n2. 存在しないアイテムIDで/api/v1/items/{item_id}エンドポイントにGETリクエストを送信', 
         '1. ステータスコード404が返される\n2. レスポンスにエラーメッセージが含まれる', 
         '低', 
         'エラーハンドリングの確認。'),
        
        ('ITEM-005', 'アイテム更新', 
         '認証済みユーザー、既存アイテム、有効な更新データ', 
         '1. 有効なJWTトークンをAuthorizationヘッダーに設定\n2. /api/v1/items/{item_id}エンドポイントにPUTリクエストを送信\n3. 有効な更新データを指定', 
         '1. ステータスコード200が返される\n2. レスポンスに更新後のデータが含まれる\n3. データベースのアイテムが更新される', 
         '中', 
         'アイテムの編集機能の確認。'),
        
        ('ITEM-006', 'アイテム削除', 
         '認証済みユーザー、既存アイテム', 
         '1. 有効なJWTトークンをAuthorizationヘッダーに設定\n2. /api/v1/items/{item_id}エンドポイントにDELETEリクエストを送信', 
         '1. ステータスコード200が返される\n2. レスポンスに成功メッセージが含まれる\n3. データベースからアイテムが削除される', 
         '中', 
         'アイテムの削除機能の確認。'),
        
        ('LLM-001', 'AIチャット', 
         '認証済みユーザー、有効なプロンプト', 
         '1. 有効なJWTトークンをAuthorizationヘッダーに設定\n2. /api/v1/llm/chatエンドポイントにPOSTリクエストを送信\n3. 有効なプロンプトとパラメータを指定', 
         '1. ステータスコード200が返される\n2. レスポンスにAIからの応答が含まれる\n3. レスポンスに使用されたモデルとトークン使用量が含まれる', 
         '高', 
         'アプリケーションの主要機能として重要。'),
        
        ('LLM-002', 'AIチャット（認証エラー）', 
         '未認証ユーザー、有効なプロンプト', 
         '1. 無効または期限切れのJWTトークンをAuthorizationヘッダーに設定\n2. /api/v1/llm/chatエンドポイントにPOSTリクエストを送信', 
         '1. ステータスコード401が返される\n2. レスポンスにエラーメッセージが含まれる', 
         '中', 
         'セキュリティ上重要な認証チェックの確認。'),
        
        ('LLM-003', 'AIチャット（無効なリクエスト）', 
         '認証済みユーザー、無効なリクエストボディ', 
         '1. 有効なJWTトークンをAuthorizationヘッダーに設定\n2. /api/v1/llm/chatエンドポイントにPOSTリクエストを送信\n3. 必須フィールドがないなど無効なリクエストボディを指定', 
         '1. ステータスコード422が返される\n2. レスポンスにバリデーションエラーメッセージが含まれる', 
         '低', 
         'リクエスト検証機能の確認。'),
        
        ('LLM-004', 'AIチャット（APIキー未設定）', 
         '認証済みユーザー、OpenAI APIキーが未設定', 
         '1. 環境変数からOpenAI APIキーを削除または無効な値に設定\n2. 有効なJWTトークンをAuthorizationヘッダーに設定\n3. /api/v1/llm/chatエンドポイントにPOSTリクエストを送信', 
         '1. ステータスコード500が返される\n2. レスポンスに構成エラーメッセージが含まれる', 
         '中', 
         '環境設定エラーハンドリングの確認。'),
        
        ('LLM-005', 'AIチャット（パフォーマンス）', 
         '認証済みユーザー、長いプロンプト', 
         '1. 有効なJWTトークンをAuthorizationヘッダーに設定\n2. /api/v1/llm/chatエンドポイントに長いプロンプトを含むPOSTリクエストを送信\n3. レスポンス時間を測定', 
         '1. ステータスコード200が返される\n2. レスポンス時間が10秒以内である', 
         '低', 
         'パフォーマンスとタイムアウト処理の確認。'),
        
        ('LLM-006', 'AI自己紹介', 
         '認証済みユーザー', 
         '1. 有効なJWTトークンをAuthorizationヘッダーに設定\n2. /api/v1/llm/masuiエンドポイントにPOSTリクエストを送信', 
         '1. ステータスコード200が返される\n2. レスポンスにAIからの自己紹介メッセージが含まれる\n3. レスポンスに使用されたモデルとトークン使用量が含まれる', 
         '中', 
         'カスタムプロンプトテンプレートの動作確認。'),
        
        ('BASE-001', 'ルートエンドポイント', 
         '任意のユーザー', 
         '1. /エンドポイントにGETリクエストを送信', 
         '1. ステータスコード200が返される\n2. レスポンスにウェルカムメッセージが含まれる', 
         '低', 
         '基本的な到達性の確認。'),
        
        ('BASE-002', 'ヘルスチェックエンドポイント', 
         '任意のユーザー', 
         '1. /healthエンドポイントにGETリクエストを送信', 
         '1. ステータスコード200が返される\n2. レスポンスにアプリケーションのステータス情報が含まれる', 
         '中', 
         'モニタリングとデプロイメントの検証に重要。'),
        
        ('UTIL-001', 'ロギングミドルウェア', 
         '任意のエンドポイントへのリクエスト', 
         '1. アプリケーションの任意のエンドポイントにリクエストを送信\n2. ログファイルとコンソール出力を確認', 
         '1. リクエスト情報（メソッド、パス、ステータスコード、処理時間）がログに記録される\n2. ログのフォーマットが適切である', 
         '中', 
         'モニタリングとデバッグに重要。'),
        
        ('UTIL-002', 'CORS設定', 
         '異なるオリジンからのリクエスト', 
         '1. 別のオリジンから/api/v1/llm/chatエンドポイントにPOSTリクエスト（プリフライトを含む）を送信', 
         '1. Access-Control-Allow-Originヘッダーが設定される\n2. プリフライトリクエストに対して適切なレスポンスが返される', 
         '低', 
         'クライアントアプリケーションとの連携に重要。'),
        
        ('UTIL-003', '設定読み込み', 
         'アプリケーション起動時', 
         '1. 環境変数を設定\n2. アプリケーションを起動\n3. ログまたはデバッグ出力で設定値を確認', 
         '1. 環境変数から設定値が正しく読み込まれる\n2. デフォルト値が適切に適用される', 
         '中', 
         '環境設定の適切な処理の確認。'),
        
        ('UTIL-004', 'JWTトークン検証', 
         '期限切れJWTトークン', 
         '1. 有効期限が切れたJWTトークンを生成\n2. そのトークンをAuthorizationヘッダーに設定\n3. 保護されたエンドポイントにリクエストを送信', 
         '1. ステータスコード401が返される\n2. レスポンスにトークン期限切れのエラーメッセージが含まれる', 
         '高', 
         'セキュリティ上重要な認証と認可の確認。'),
    ]
    
    for row_num, test_case in enumerate(test_cases, 2):  # ヘッダー行の後から開始
        for col_num, value in enumerate(test_case, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = normal_font
            cell.alignment = normal_alignment
            cell.border = thin_border
    
    save_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "DOC", "単体テスト")
    os.makedirs(save_dir, exist_ok=True)
    
    filename = os.path.join(save_dir, "単体テスト仕様書.xlsx")
    wb.save(filename)
    print(f"単体テスト仕様書を作成しました: {filename}")
    
    return filename

if __name__ == "__main__":
    create_unit_test_spec()
