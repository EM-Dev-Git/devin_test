"""
単体テスト実行とエビデンス収集スクリプト

このスクリプトは単体テストを実行し、テスト結果のエビデンスを収集します。
テスト結果はExcelレポートとして保存され、ログファイルとスクリーンショットも
エビデンスとして保存されます。
"""
import os
import sys
import time
import datetime
import subprocess
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import shutil
import json
from PIL import Image, ImageDraw, ImageFont
import pytest

EVIDENCE_DIR = os.path.join("DOC", "単体テスト", "エビデンス")
os.makedirs(EVIDENCE_DIR, exist_ok=True)

now = datetime.datetime.now()
timestamp = now.strftime("%Y%m%d_%H%M%S")

excel_path = os.path.join(EVIDENCE_DIR, f"単体テスト結果_{timestamp}.xlsx")

LOG_DIR = os.path.join(EVIDENCE_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)

SCREENSHOT_DIR = os.path.join(EVIDENCE_DIR, "screenshots")
os.makedirs(SCREENSHOT_DIR, exist_ok=True)

def create_test_report(test_results):
    """
    テスト結果をExcelレポートとして保存する
    
    Args:
        test_results: テスト結果のリスト
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "単体テスト結果"
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 20
    
    header_font = Font(name='メイリオ', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type='solid', fgColor="4472C4")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    normal_font = Font(name='メイリオ', size=11)
    normal_alignment = Alignment(vertical='top', wrap_text=True)
    
    success_fill = PatternFill(fill_type='solid', fgColor="C6EFCE")
    failure_fill = PatternFill(fill_type='solid', fgColor="FFC7CE")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["テストID", "テスト対象", "ステータス", "実行時間(秒)", "詳細", "実行日時"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_num, result in enumerate(test_results, 2):
        cell = ws.cell(row=row_num, column=1)
        cell.value = result["test_id"]
        cell.font = normal_font
        cell.alignment = normal_alignment
        cell.border = thin_border
        
        cell = ws.cell(row=row_num, column=2)
        cell.value = result["test_name"]
        cell.font = normal_font
        cell.alignment = normal_alignment
        cell.border = thin_border
        
        cell = ws.cell(row=row_num, column=3)
        cell.value = result["status"]
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        if result["status"] == "成功":
            cell.fill = success_fill
        else:
            cell.fill = failure_fill
        
        cell = ws.cell(row=row_num, column=4)
        cell.value = round(result["execution_time"], 2)
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = thin_border
        
        cell = ws.cell(row=row_num, column=5)
        cell.value = result["details"]
        cell.font = normal_font
        cell.alignment = normal_alignment
        cell.border = thin_border
        
        cell = ws.cell(row=row_num, column=6)
        cell.value = result["timestamp"]
        cell.font = normal_font
        cell.alignment = normal_alignment
        cell.border = thin_border
    
    ws_summary = wb.create_sheet(title="サマリー")
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15
    
    ws_summary.cell(row=1, column=1).value = "項目"
    ws_summary.cell(row=1, column=2).value = "値"
    ws_summary.cell(row=1, column=1).font = header_font
    ws_summary.cell(row=1, column=1).fill = header_fill
    ws_summary.cell(row=1, column=1).alignment = header_alignment
    ws_summary.cell(row=1, column=1).border = thin_border
    ws_summary.cell(row=1, column=2).font = header_font
    ws_summary.cell(row=1, column=2).fill = header_fill
    ws_summary.cell(row=1, column=2).alignment = header_alignment
    ws_summary.cell(row=1, column=2).border = thin_border
    
    success_count = sum(1 for result in test_results if result["status"] == "成功")
    failure_count = len(test_results) - success_count
    total_count = len(test_results)
    success_rate = (success_count / total_count) * 100 if total_count > 0 else 0
    
    summary_data = [
        ("テスト実行日時", now.strftime("%Y-%m-%d %H:%M:%S")),
        ("合計テスト数", total_count),
        ("成功数", success_count),
        ("失敗数", failure_count),
        ("成功率", f"{success_rate:.1f}%"),
        ("合計実行時間", f"{sum(result['execution_time'] for result in test_results):.2f}秒")
    ]
    
    for row_num, (item, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row_num, column=1).value = item
        ws_summary.cell(row=row_num, column=1).font = normal_font
        ws_summary.cell(row=row_num, column=1).alignment = normal_alignment
        ws_summary.cell(row=row_num, column=1).border = thin_border
        
        ws_summary.cell(row=row_num, column=2).value = value
        ws_summary.cell(row=row_num, column=2).font = normal_font
        ws_summary.cell(row=row_num, column=2).alignment = normal_alignment
        ws_summary.cell(row=row_num, column=2).border = thin_border
    
    wb.save(excel_path)
    print(f"テスト結果レポートを保存しました: {excel_path}")
    
    return excel_path

def run_tests():
    """
    単体テストを実行し、結果を収集する
    
    Returns:
        テスト結果のリスト
    """
    print("単体テストを実行中...")
    
    os.makedirs("tests/temp", exist_ok=True)
    
    log_file = f"tests/temp/test_log_{timestamp}.log"
    
    test_command = [
        "python", "-m", "pytest",
        "tests",
        "-v"
    ]
    
    with open(log_file, "w") as f:
        subprocess.run(test_command, stdout=f, stderr=subprocess.STDOUT)
    
    test_results = []
    
    with open(log_file, "r") as f:
        log_content = f.read()
        
        import re
        test_pattern = r"(test_[a-zA-Z0-9_]+) (PASSED|FAILED|SKIPPED|XFAILED|XPASSED)"
        test_matches = re.findall(test_pattern, log_content)
        
        for test_name, test_result in test_matches:
            test_id_pattern = rf"{test_name}.*?\n.*?((?:BASE|AUTH|ITEM|LLM|UTIL)-\d+)"
            test_id_match = re.search(test_id_pattern, log_content, re.DOTALL)
            test_id = test_id_match.group(1) if test_id_match else "UNKNOWN"
            
            time_pattern = rf"{test_name}.*?(\d+\.\d+)s"
            time_match = re.search(time_pattern, log_content)
            execution_time = float(time_match.group(1)) if time_match else 0.0
            
            details_pattern = rf"{test_name}.*?details': '(.*?)'"
            details_match = re.search(details_pattern, log_content, re.DOTALL)
            details = details_match.group(1) if details_match else ""
            
            test_name_formatted = test_name.replace("test_", "").replace("_", " ").title()
            
            test_results.append({
                "test_id": test_id,
                "test_name": test_name_formatted,
                "status": "成功" if test_result == "PASSED" else "失敗",
                "execution_time": execution_time,
                "details": details if details else (
                    "テストが正常に完了しました。" if test_result == "PASSED" else "テストが失敗しました。"
                ),
                "timestamp": now.strftime("%Y-%m-%d %H:%M:%S")
            })
    
    shutil.copy(log_file, os.path.join(LOG_DIR, f"test_log_{timestamp}.log"))
    
    create_test_summary_image(test_results)
    
    return test_results

def create_test_summary_image(test_results):
    """
    テスト結果のサマリースクリーンショットを作成する
    
    Args:
        test_results: テスト結果のリスト
    """
    success_count = sum(1 for result in test_results if result["status"] == "成功")
    failure_count = len(test_results) - success_count
    total_count = len(test_results)
    success_rate = (success_count / total_count) * 100 if total_count > 0 else 0
    
    width, height = 800, 600
    image = Image.new("RGB", (width, height), (255, 255, 255))
    draw = ImageDraw.Draw(image)
    
    draw.text((50, 50), "単体テスト結果サマリー", fill=(0, 0, 0))
    draw.text((50, 100), f"実行日時: {now.strftime('%Y-%m-%d %H:%M:%S')}", fill=(0, 0, 0))
    draw.text((50, 150), f"合計テスト数: {total_count}", fill=(0, 0, 0))
    draw.text((50, 200), f"成功数: {success_count}", fill=(0, 0, 128))
    draw.text((50, 250), f"失敗数: {failure_count}", fill=(128, 0, 0))
    draw.text((50, 300), f"成功率: {success_rate:.1f}%", fill=(0, 0, 0))
    
    categories = {
        "BASE": "基本エンドポイント",
        "AUTH": "認証機能",
        "ITEM": "アイテムCRUD",
        "LLM": "OpenAI連携",
        "UTIL": "ユーティリティ"
    }
    
    y_pos = 350
    for prefix, category_name in categories.items():
        category_tests = [r for r in test_results if r["test_id"].startswith(prefix)]
        if category_tests:
            category_success = sum(1 for r in category_tests if r["status"] == "成功")
            category_rate = (category_success / len(category_tests)) * 100
            draw.text((50, y_pos), f"{category_name}: {category_rate:.1f}% ({category_success}/{len(category_tests)})", fill=(0, 0, 0))
            y_pos += 50
    
    image_path = os.path.join(SCREENSHOT_DIR, f"test_summary_{timestamp}.png")
    image.save(image_path)
    print(f"テスト結果サマリースクリーンショットを保存しました: {image_path}")

if __name__ == "__main__":
    test_results = run_tests()
    
    report_path = create_test_report(test_results)
    
    success_count = sum(1 for result in test_results if result["status"] == "成功")
    total_count = len(test_results)
    success_rate = (success_count / total_count) * 100 if total_count > 0 else 0
    
    print(f"単体テスト実行完了。テスト成功率: {success_rate:.1f}% ({success_count}/{total_count})")
    print(f"テスト結果レポート: {report_path}")
    print(f"テストログ: {os.path.join(LOG_DIR, f'test_log_{timestamp}.log')}")
    print(f"テスト結果サマリー画像: {os.path.join(SCREENSHOT_DIR, f'test_summary_{timestamp}.png')}")
