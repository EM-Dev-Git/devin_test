"""
アーキテクチャ図の基本的なExcel形式を生成するスクリプト

このスクリプトはopenpyxlライブラリを使用して、
システムアーキテクチャ図の基本的なExcel形式を生成します。
"""
import os
import sys
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import requests
from io import BytesIO
from PIL import Image as PILImage

def download_icon(url, size=(48, 48)):
    """
    URLから画像をダウンロードし、指定されたサイズにリサイズする
    
    Args:
        url: 画像のURL
        size: リサイズするサイズ（幅, 高さ）
        
    Returns:
        BytesIO: 画像データ
    """
    try:
        response = requests.get(url, stream=True)
        response.raise_for_status()
        
        img = PILImage.open(BytesIO(response.content))
        img = img.convert("RGBA")
        img = img.resize(size)
        
        buffered = BytesIO()
        img.save(buffered, format="PNG")
        buffered.seek(0)
        
        return buffered
    except Exception as e:
        print(f"Error downloading image from {url}: {str(e)}")
        return None

def create_excel_architecture(output_file):
    """
    アーキテクチャ図の基本的なExcel形式を生成する
    
    Args:
        output_file: 出力Excelファイルのパス
        
    Returns:
        bool: 生成が成功した場合はTrue、それ以外はFalse
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "システムアーキテクチャ図"
        
        for col in range(1, 30):
            ws.column_dimensions[get_column_letter(col)].width = 3
        
        for row in range(1, 50):
            ws.row_dimensions[row].height = 15
        
        ws['A1'] = "EM_test_project システムアーキテクチャ図"
        ws['A1'].font = Font(name='Arial', size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws['B3'] = "Windows ホストシステム"
        ws['B3'].font = Font(name='Arial', size=12, bold=True)
        ws['B3'].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        
        ws['D5'] = "WSL (Windows Subsystem for Linux)"
        ws['D5'].font = Font(name='Arial', size=11, bold=True)
        ws['D5'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        ws['F7'] = "FastAPI アプリケーション"
        ws['F7'].font = Font(name='Arial', size=10, bold=True)
        ws['F7'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        ws['G9'] = "FastAPI コア"
        ws['G9'].font = Font(name='Arial', size=9)
        ws['G9'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        ws['G11'] = "認証モジュール"
        ws['G11'].font = Font(name='Arial', size=9)
        ws['G11'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        ws['G13'] = "OpenAI クライアント"
        ws['G13'].font = Font(name='Arial', size=9)
        ws['G13'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        ws['P9'] = "SQLite データベース"
        ws['P9'].font = Font(name='Arial', size=10)
        ws['P9'].fill = PatternFill(start_color="F8CBCE", end_color="F8CBCE", fill_type="solid")
        
        ws['Y7'] = "Microsoft Azure"
        ws['Y7'].font = Font(name='Arial', size=10, bold=True)
        ws['Y7'].fill = PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid")
        
        ws['Z10'] = "Azure OpenAI Service"
        ws['Z10'].font = Font(name='Arial', size=9)
        ws['Z10'].fill = PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid")
        
        ws['D22'] = "クライアント"
        ws['D22'].font = Font(name='Arial', size=9)
        
        ws['M11'] = "SQLAlchemy ORM"
        ws['M11'].font = Font(name='Arial', size=8)
        
        ws['Q17'] = "HTTPS / REST API"
        ws['Q17'].font = Font(name='Arial', size=8)
        
        ws['G18'] = "HTTP / REST API"
        ws['G18'].font = Font(name='Arial', size=8)
        
        ws['Y15'] = "凡例"
        ws['Y15'].font = Font(name='Arial', size=10, bold=True)
        
        legend_items = [
            "Windows ホストシステム: 物理マシン",
            "WSL: Linux 仮想環境",
            "FastAPI: Python Webアプリケーション",
            "Azure OpenAI: マイクロソフトのAIサービス"
        ]
        
        for i, item in enumerate(legend_items):
            ws[f"Y{16+i}"] = item
            ws[f"Y{16+i}"].font = Font(name='Arial', size=8)
        
        icon_urls = {
            "windows": "https://cdn-icons-png.flaticon.com/512/888/888882.png",
            "linux": "https://cdn-icons-png.flaticon.com/512/6124/6124995.png",
            "fastapi": "https://fastapi.tiangolo.com/img/logo-margin/logo-teal.png",
            "database": "https://cdn-icons-png.flaticon.com/512/148/148825.png",
            "azure": "https://cdn-icons-png.flaticon.com/512/873/873107.png",
            "openai": "https://upload.wikimedia.org/wikipedia/commons/thumb/0/04/ChatGPT_logo.svg/512px-ChatGPT_logo.svg.png",
            "client": "https://cdn-icons-png.flaticon.com/512/2521/2521826.png"
        }
        
        icon_positions = {
            "windows": "B3",
            "linux": "D5",
            "fastapi": "F7",
            "database": "P9",
            "azure": "Y7",
            "openai": "Z10",
            "client": "D22"
        }
        
        icons_dir = "icons"
        if not os.path.exists(icons_dir):
            os.makedirs(icons_dir)
        
        for name, url in icon_urls.items():
            try:
                icon_data = download_icon(url)
                if icon_data:
                    icon_path = os.path.join(icons_dir, f"{name}.png")
                    with open(icon_path, "wb") as f:
                        f.write(icon_data.getvalue())
                    
                    img = Image(icon_path)
                    img.width = 30
                    img.height = 30
                    
                    cell_position = icon_positions[name]
                    ws.add_image(img, cell_position)
                    
                    print(f"Added {name} icon to Excel")
            except Exception as e:
                print(f"Error adding {name} icon to Excel: {str(e)}")
        
        wb.save(output_file)
        print(f"Successfully created Excel architecture diagram at {output_file}")
        return True
        
    except Exception as e:
        print(f"Error creating Excel architecture diagram: {str(e)}")
        return False

def main():
    """
    メイン関数
    """
    output_file = "DOC/アーキテクチャ/システムアーキテクチャ図.xlsx"
    
    output_dir = os.path.dirname(output_file)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    success = create_excel_architecture(output_file)
    
    if success:
        print(f"Excel generation successful. Excel file saved to {output_file}")
        return 0
    else:
        print("Excel generation failed")
        return 1

if __name__ == "__main__":
    sys.exit(main())
