"""
コーディング規約のExcelドキュメントを生成するスクリプト

このスクリプトはopenpyxlライブラリを使用して、
FastAPIプロジェクトのコーディング規約をExcel形式で生成します。
"""
import os
import sys
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

def create_coding_standards_doc(output_file):
    """
    コーディング規約のExcelドキュメントを生成する
    
    Args:
        output_file: 出力Excelファイルのパス
        
    Returns:
        bool: 生成が成功した場合はTrue、それ以外はFalse
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "コーディング規約"
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 60
        
        ws['B1'] = "FastAPI コーディング規約"
        ws['B1'].font = Font(name='Yu Gothic', size=16, bold=True)
        ws['B1'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws['B2'] = "目的"
        ws['B2'].font = Font(name='Yu Gothic', size=12, bold=True)
        ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws['C2'] = "このドキュメントはFastAPIプロジェクトのコーディング規約を定義し、チーム開発やコードの一貫性を保つことを目的としています。"
        ws['C2'].font = Font(name='Yu Gothic', size=11)
        ws['C2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        row = 4
        ws[f'B{row}'] = "1. プロジェクト構成"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=12, bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
        ws[f'C{row}'] = """
app/
├── api/              # ルーティング関連
│   ├── v1/
│   │   ├── endpoints/
│   │   │   └── user.py
│   │   └── router.py
├── core/             # 設定や初期化処理
│   ├── config.py
│   └── security.py
├── models/           # PydanticモデルやORMモデル
│   └── user.py
├── services/         # ビジネスロジック
│   └── user_service.py
├── db/               # DB接続やセッション管理
│   ├── base.py
│   └── session.py
├── main.py           # アプリケーションのエントリーポイント
└── dependencies/     # 共通依存関係
        """
        ws[f'C{row}'].font = Font(name='Consolas', size=10)
        ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = 200
        
        row += 2
        ws[f'B{row}'] = "2. 命名規則"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=12, bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
        ws[f'B{row}'] = "ファイル名"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'C{row}'] = "snake_case（例: user_service.py）"
        ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
        
        row += 1
        ws[f'B{row}'] = "クラス名"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'C{row}'] = "PascalCase（例: UserService）"
        ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
        
        row += 1
        ws[f'B{row}'] = "関数・変数名"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'C{row}'] = "snake_case（例: get_user_by_id）"
        ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
        
        row += 1
        ws[f'B{row}'] = "定数"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'C{row}'] = "UPPER_CASE（例: MAX_CONNECTIONS）"
        ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
        
        row += 2
        ws[f'B{row}'] = "3. 型ヒントとドキュメンテーション"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=12, bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
        code_example = """
from typing import List
from fastapi import APIRouter

router = APIRouter()

@router.get("/users", response_model=List[UserResponse])
def get_users() -> List[UserResponse]:
    \"\"\"
    ユーザー一覧を取得するエンドポイント
    \"\"\"
    return user_service.get_all_users()
        """
        ws[f'C{row}'] = code_example
        ws[f'C{row}'].font = Font(name='Consolas', size=10)
        ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = 150
        
        row += 2
        ws[f'B{row}'] = "4. Pydanticモデルの使い方"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=12, bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
        ws[f'C{row}'] = """
from pydantic import BaseModel, EmailStr

class UserCreate(BaseModel):
    email: EmailStr
    password: str

    class Config:
        orm_mode = True
        """
        ws[f'C{row}'].font = Font(name='Consolas', size=10)
        ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = 120
        
        row += 2
        ws[f'B{row}'] = "5. 依存関係の注入"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=12, bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
        ws[f'C{row}'] = """
from fastapi import Depends
from sqlalchemy.orm import Session
from app.db.session import get_db

@router.post("/users")
def create_user(user: UserCreate, db: Session = Depends(get_db)):
    return user_service.create_user(db, user)
        """
        ws[f'C{row}'].font = Font(name='Consolas', size=10)
        ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = 120
        
        row += 2
        ws[f'B{row}'] = "6. 設定ファイル"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=12, bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
        ws[f'C{row}'] = """
from pydantic import BaseSettings

class Settings(BaseSettings):
    app_name: str = "My FastAPI App"
    database_url: str

    class Config:
        env_file = ".env"

settings = Settings()
        """
        ws[f'C{row}'].font = Font(name='Consolas', size=10)
        ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = 150
        
        row += 2
        ws[f'B{row}'] = "7. ログ出力の統一"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=12, bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
        ws[f'C{row}'] = """
import logging

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
        """
        ws[f'C{row}'].font = Font(name='Consolas', size=10)
        ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = 80
        
        row += 2
        ws[f'B{row}'] = "8. テストの命名と構成"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=12, bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
        ws[f'B{row}'] = "テストの配置"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'C{row}'] = "`tests/` ディレクトリに配置"
        ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
        
        row += 1
        ws[f'B{row}'] = "テスト関数の命名"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'C{row}'] = "テスト関数は `test_` で始める"
        ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
        
        row += 1
        ws[f'B{row}'] = "テストフレームワーク"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'C{row}'] = "`pytest` を使用"
        ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
        
        row += 1
        ws[f'C{row}'] = """
def test_create_user(client):
    response = client.post("/users", json={"email": "test@example.com", "password": "123456"})
    assert response.status_code == 200
        """
        ws[f'C{row}'].font = Font(name='Consolas', size=10)
        ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = 80
        
        row += 2
        ws[f'B{row}'] = "9. 推奨ツール"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=12, bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
        ws[f'B{row}'] = "コードフォーマッター"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'C{row}'] = "Black"
        ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
        
        row += 1
        ws[f'B{row}'] = "インポート整理"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'C{row}'] = "isort"
        ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
        
        row += 1
        ws[f'B{row}'] = "静的型チェック"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'C{row}'] = "mypy"
        ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
        
        row += 1
        ws[f'B{row}'] = "リンター"
        ws[f'B{row}'].font = Font(name='Yu Gothic', size=11, bold=True)
        ws[f'C{row}'] = "ruff"
        ws[f'C{row}'].font = Font(name='Yu Gothic', size=11)
        
        wb.save(output_file)
        print(f"Successfully created coding standards document at {output_file}")
        return True
        
    except Exception as e:
        print(f"Error creating coding standards document: {str(e)}")
        return False

def main():
    """
    メイン関数
    """
    coding_standards_dir = "DOC/コーディング規約"
    if not os.path.exists(coding_standards_dir):
        os.makedirs(coding_standards_dir)
        
    readme_path = os.path.join(coding_standards_dir, "README.md")
    with open(readme_path, "w", encoding="utf-8") as f:
        f.write("""# コーディング規約

このディレクトリにはFastAPIプロジェクトのコーディング規約に関するドキュメントが含まれています。


- `コーディング規約.xlsx`: FastAPIプロジェクトのコーディング規約を定義したExcelファイル
""")
    
    output_file = os.path.join(coding_standards_dir, "コーディング規約.xlsx")
    success = create_coding_standards_doc(output_file)
    
    if success:
        print(f"Excel generation successful. Excel file saved to {output_file}")
        return 0
    else:
        print("Excel generation failed")
        return 1

if __name__ == "__main__":
    sys.exit(main())
