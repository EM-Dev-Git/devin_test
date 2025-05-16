"""
Excel ファイルの整合性を検証するスクリプト

このスクリプトは、作成したExcelファイルが正常に開けるかどうかを確認します。
ファイルが破損している場合は、エラーメッセージを表示します。
"""
import os
import sys
from openpyxl import load_workbook

def verify_excel_file(file_path):
    """
    Excelファイルが正常に開けるかどうかを確認する
    
    Args:
        file_path: 検証するExcelファイルのパス
        
    Returns:
        bool: ファイルが正常に開ける場合はTrue、それ以外はFalse
    """
    try:
        print(f"検証中: {file_path}")
        if not os.path.exists(file_path):
            print(f"エラー: ファイルが存在しません: {file_path}")
            return False
            
        wb = load_workbook(file_path)
        
        sheet_names = wb.sheetnames
        print(f"  シート数: {len(sheet_names)}")
        print(f"  シート名: {', '.join(sheet_names)}")
        
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            print(f"  シート '{sheet_name}': {sheet.max_row} 行 x {sheet.max_column} 列")
            
            if sheet.max_row > 0 and sheet.max_column > 0:
                sample_cell = sheet.cell(row=1, column=1).value
                print(f"  最初のセルの内容: {sample_cell}")
        
        print(f"結果: ファイルは正常です: {file_path}")
        return True
    except Exception as e:
        print(f"エラー: ファイルが破損しています: {file_path}")
        print(f"エラー詳細: {str(e)}")
        return False

def main():
    """
    すべてのExcelファイルを検証する
    """
    excel_files = [
        "DOC/要件定義/OpenAI連携機能_要件定義書.xlsx",
        "DOC/外部設計/OpenAI連携機能_外部設計書.xlsx",
        "DOC/内部設定/OpenAI連携機能_内部設計書.xlsx",
        "DOC/単体テスト/OpenAI連携機能_プログラム詳細設計書.xlsx"
    ]
    
    all_valid = True
    for file_path in excel_files:
        if not verify_excel_file(file_path):
            all_valid = False
            
    if all_valid:
        print("\nすべてのExcelファイルは正常です。")
        return 0
    else:
        print("\n一部のExcelファイルが破損しています。再作成が必要です。")
        return 1

if __name__ == "__main__":
    sys.exit(main())
